#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gestiq v2.1
Automatización IMAGINE + GUARDIAN — Interfaz moderna
Requiere: pip install playwright openpyxl customtkinter
          playwright install chromium
"""

import sys, os, re, json, base64, threading, asyncio, webbrowser
from datetime import datetime
from urllib.parse import unquote
import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

try:
    import customtkinter as ctk
    ctk.set_appearance_mode("dark")
    ctk.set_default_color_theme("green")
    HAS_CTK = True
except ImportError:
    HAS_CTK = False
    print("CustomTkinter no encontrado. Ejecuta: pip install customtkinter")
    sys.exit(1)

# ── Tipografía del sistema (SF Pro en Mac, Segoe UI en Windows) ──────────────
MONO_FAM = "Menlo" if sys.platform == "darwin" else "Consolas"
FONT_FAM = None   # se resuelve al crear la ventana (requiere Tk activo)

def _init_fonts():
    global FONT_FAM
    if FONT_FAM:
        return
    try:
        import tkinter.font as tkfont
        fams = set(tkfont.families())
    except Exception:
        fams = set()
    # En Mac, SF Pro reporta mal las métricas en Tk y corta los trazos de g/q/y;
    # Helvetica Neue se ve casi igual y no recorta.
    prefer = (("Helvetica Neue", "SF Pro Text", "SF Pro Display")
              if sys.platform == "darwin" else ("Segoe UI", "Calibri"))
    FONT_FAM = next((f for f in prefer if f in fams), prefer[-1])

def F(size, bold=False):
    """Fuente consistente en toda la app."""
    # Tk 9 en Mac recorta los trazos de g/q/y/p con tamaño 11: mínimo 12.
    if sys.platform == "darwin" and size < 12:
        size = 12
    return ctk.CTkFont(family=FONT_FAM, size=size,
                       weight="bold" if bold else "normal")

# ── Logos de módulos (PNG junto al script o dentro del ejecutable) ───────────
def _res_path(name):
    base = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, name)

def _load_logo(name, size=22):
    """CTkImage desde PNG; None si falta el archivo o Pillow."""
    try:
        from PIL import Image
        img = Image.open(_res_path(name))
        # misma altura visual para todos y lienzo de ancho fijo, centrado
        H = 128
        w = max(1, round(img.width * H / img.height))
        img = img.resize((w, H), Image.LANCZOS)
        W = round(H * 1.2)
        sq = Image.new("RGBA", (W, H), (0, 0, 0, 0))
        sq.paste(img, (0, 0))                    # alineados al borde izquierdo
        return ctk.CTkImage(light_image=sq, dark_image=sq,
                            size=(round(size * 1.2), size))
    except Exception:
        return None

try:
    from playwright.async_api import async_playwright, TimeoutError as PWTimeout
    HAVE_PW = True
except ImportError:
    HAVE_PW = False

try:
    import licencia
except Exception:
    licencia = None

# Web de registro y pago
REGISTRO_URL = "https://toxicamazonas-svg.github.io/gestiq/cuenta.html"

# ── Paleta sobria (tuplas CTK: primer valor = claro, segundo = oscuro) ───────
# 4 colores base: fondo neutro, acento verde Bolívar, gris secundario, rojo errores
BG     = ("#F4F4F6", "#1E1E2E")   # fondo neutro
CARD   = ("#FFFFFF", "#27273A")   # tarjetas
CARD2  = ("#EDEDF2", "#313147")   # campos / zonas internas
BORDER = ("#DEDEE6", "#3B3B54")
TX     = ("#1B1B2B", "#EDEDF5")   # texto principal
TM     = ("#6E7285", "#9A9FB4")   # texto secundario (gris)

ACCENT   = "#00B050"              # verde Bolívar (acento corporativo)
ACCENT_H = "#009A46"              # hover del acento
ERR      = "#E5484D"              # rojo, solo para errores
ERR_H    = "#C53A3F"

# Consola de registro (fondo oscuro fijo en ambos temas)
LOG_BG, LOG_FG = "#14141F", "#E2E6F0"
OK_C = "#3DDC84"   # ✓ verde
ER_C = "#FF6B6B"   # ✗ rojo
WA_C = "#FFA94D"   # ⚠ naranja
OR_C = "#FF8A3D"   # ◆ naranja oscuro (sin resultado)

# ── Estilos Excel (mismos colores/formato del archivo original) ─────────────
def xfill(rgb): return PatternFill(start_color=rgb, end_color=rgb, fill_type='solid')
F_GREEN   = xfill('00B050')   # verde sólido (APROBADO)
F_GREENLT = xfill('92D050')   # verde claro (observaciones autorizadas)
F_RED     = xfill('FF0000')   # rojo (devuelto / rechazado)
F_ORANGE  = xfill('FFC000')   # naranja (no requiere / errado)
F_YELLOW  = xfill('FFFF00')   # amarillo (pendiente)

AL_CELL = Alignment(horizontal='center', vertical='center', wrap_text=True)

def style_cell(cell, fill=None, white=False):
    """Aplica el formato del Excel original: Calibri 7 negrita, centrado, ajustar texto."""
    if fill: cell.fill = fill
    cell.font      = Font(name='Calibri', size=7, bold=True,
                          color='FFFFFF' if white else '000000')
    cell.alignment = AL_CELL


# ════════════════════════════════════════════════════════════════════════════
#  Ventana principal
# ════════════════════════════════════════════════════════════════════════════
class App(ctk.CTk):
    VERSION = "v2.1"

    def __init__(self):
        super().__init__()
        _init_fonts()
        self.title("Gestiq")
        self.geometry("1020x720")
        self.minsize(900, 650)
        self.configure(fg_color=BG)
        self._build()
        self.protocol("WM_DELETE_WINDOW", self._on_close)
        self._lic = None            # sesión de licencia activa
        self._lock = None           # overlay de login/bloqueo
        self._hb = None             # heartbeat de licencia
        self._plan = "completo"     # imagine | guardian | completo
        self.after(80, self._lic_iniciar)
        if not HAVE_PW:
            messagebox.showwarning(
                "Falta un componente",
                "Playwright no está instalado, así que el bot no podrá abrir "
                "el navegador.\n\n"
                "Abre la Terminal y ejecuta:\n\n"
                "  pip3 install playwright openpyxl customtkinter\n"
                "  playwright install chromium\n\n"
                "Luego vuelve a abrir el programa.")

    def _toggle_theme(self):
        dark = bool(self.sw_theme.get())
        # CTk hace un volcado de pantalla por cada widget al cambiar el tema
        # (update_idletasks en cada uno), y eso se ve como cambio progresivo.
        # Se desactiva ese volcado durante el cambio y se hace uno solo al final.
        orig = tk.Misc.update_idletasks
        tk.Misc.update_idletasks = lambda *_: None
        try:
            ctk.set_appearance_mode("dark" if dark else "light")
        finally:
            tk.Misc.update_idletasks = orig
        self.sw_theme.configure(text="Oscuro" if dark else "Claro")
        self.update_idletasks()

    def _build(self):
        self.columnconfigure(1, weight=1)
        self.rowconfigure(1, weight=1)

        # ── Encabezado ───────────────────────────────────────────────────────
        hdr = ctk.CTkFrame(self, fg_color=CARD, corner_radius=0, height=64)
        hdr.grid(row=0, column=0, columnspan=2, sticky="ew")
        hdr.pack_propagate(False)

        logo = ctk.CTkFrame(hdr, fg_color=ACCENT, corner_radius=18,
                            width=36, height=36)
        logo.pack(side="left", padx=(20, 12), pady=14)
        logo.pack_propagate(False)
        ctk.CTkLabel(logo, text="IP", font=F(13, True),
                     text_color="#FFFFFF").place(relx=.5, rely=.5, anchor="center")

        ctk.CTkLabel(hdr, text="Gestiq", font=F(19, True),
                     text_color=TX).pack(side="left")

        theme_box = ctk.CTkFrame(hdr, fg_color=CARD2, corner_radius=10)
        theme_box.pack(side="right", padx=20, pady=12)
        ctk.CTkLabel(theme_box, text="Modo de visualización", font=F(11),
                     text_color=TM).pack(side="left", padx=(12, 6))
        self.sw_theme = ctk.CTkSwitch(theme_box, text="Oscuro", font=F(11),
                                      text_color=TM, progress_color=ACCENT,
                                      command=self._toggle_theme)
        self.sw_theme.pack(side="left", padx=(0, 10), pady=6)
        self.sw_theme.select()

        # Panel de sesión (se muestra solo con sesión iniciada)
        self.sess_box = ctk.CTkFrame(hdr, fg_color=CARD2, corner_radius=10)
        self.lbl_sess = ctk.CTkLabel(self.sess_box, text="", font=F(11),
                                     text_color=TM)
        self.lbl_sess.pack(side="left", padx=(12, 8), pady=6)
        ctk.CTkButton(self.sess_box, text="Cerrar sesión", font=F(11),
                      width=106, height=28, fg_color="transparent",
                      border_width=1, border_color=BORDER, text_color=TX,
                      hover_color=CARD,
                      command=self._lic_cerrar_sesion).pack(side="left",
                                                            padx=(0, 10), pady=6)

        # ── Barra lateral ────────────────────────────────────────────────────
        side = ctk.CTkFrame(self, fg_color=CARD, corner_radius=0, width=190)
        side.grid(row=1, column=0, sticky="nsw")
        side.pack_propagate(False)

        ctk.CTkLabel(side, text="MÓDULOS", font=F(11, True), text_color=TM,
                     anchor="w").pack(fill="x", padx=20, pady=(20, 6))

        self._nav = {}
        for key, emoji, label in (("imagine", "🔍", "IMAGINE"),
                                  ("guardian", "🛡", "GUARDIÁN")):
            logo = _load_logo(f"logo_{key}.png")
            txt = f"  {label}" if logo else f"{emoji}   {label}"
            b = ctk.CTkButton(side, text=txt, image=logo, compound="left",
                              anchor="w", height=40,
                              corner_radius=10, font=F(13),
                              fg_color="transparent", hover_color=CARD2,
                              text_color=TM,
                              command=lambda k=key: self._show(k))
            b.pack(fill="x", padx=12, pady=3)
            self._nav[key] = b

        ctk.CTkLabel(side, text="Consulta casos en los\nportales y registra los\n"
                                "resultados en tu Excel.",
                     font=F(11), text_color=TM, justify="left",
                     anchor="w").pack(side="bottom", fill="x", padx=20, pady=18)

        # ── Contenido ────────────────────────────────────────────────────────
        content = ctk.CTkFrame(self, fg_color="transparent")
        content.grid(row=1, column=1, sticky="nsew", padx=24, pady=20)
        content.columnconfigure(0, weight=1)
        content.rowconfigure(0, weight=1)

        self.imagine_tab  = ImagineTab(content, self)
        self.guardian_tab = GuardianTab(content, self)
        for t in (self.imagine_tab, self.guardian_tab):
            t.grid(row=0, column=0, sticky="nsew")
        self._show("imagine")

    # Módulos permitidos por plan
    _PLAN_MODS = {"imagine": ("imagine",), "guardian": ("guardian",),
                  "completo": ("imagine", "guardian")}

    def _mods_permitidos(self):
        return self._PLAN_MODS.get(self._plan, ("imagine", "guardian"))

    def _show(self, key):
        if key not in self._mods_permitidos():
            messagebox.showinfo(
                "Módulo no incluido",
                "Tu plan actual no incluye este módulo.\n"
                "Puedes ampliarlo en la página de tu cuenta.")
            return
        self._mod_actual = key
        for k, b in self._nav.items():
            active = (k == key)
            b.configure(fg_color=ACCENT if active else "transparent",
                        text_color="#FFFFFF" if active else TM,
                        hover_color=ACCENT_H if active else CARD2)
        (self.imagine_tab if key == "imagine" else self.guardian_tab).lift()

    def _set_plan(self, plan):
        """Actualiza el plan y habilita/deshabilita módulos en la sidebar."""
        plan = str(plan or "completo").lower()
        if plan not in self._PLAN_MODS:      # valores antiguos (p.ej. 'mensual')
            plan = "completo"
        self._plan = plan
        mods = self._mods_permitidos()
        for k, b in self._nav.items():
            b.configure(state="normal" if k in mods else "disabled")
        if getattr(self, "_mod_actual", "imagine") not in mods:
            self._show(mods[0])

    def _on_close(self):
        if any(t._running for t in (self.imagine_tab, self.guardian_tab)):
            if not messagebox.askyesno(
                    "Proceso en curso",
                    "Hay una consulta en curso. Si cierras ahora se detendrá y "
                    "podrías perder resultados no guardados.\n\n"
                    "¿Cerrar de todas formas?"):
                return
        self.destroy()

    # ── Licencias (Supabase) — todo se valida en el servidor ────────────────
    def _lic_iniciar(self):
        if licencia is None or not licencia.configurado():
            self._lock_mostrar("bloqueado",
                               "Error interno de licencias. Reinstala la aplicación "
                               "o contacta soporte.")
            return
        self._lock_mostrar("cargando")
        def work():
            try:
                s = licencia.restaurar_sesion()
                if s is None:
                    self.after(0, lambda: self._lock_mostrar("login")); return
                r = licencia.verificar(s)
                if r.get("ok"):
                    self.after(0, lambda r=r: self._lic_ok(s, r))
                else:
                    self.after(0, lambda m=licencia.motivo(r):
                               self._lock_mostrar("bloqueado", m))
            except Exception as e:
                self.after(0, lambda m=str(e):
                           self._lock_mostrar("bloqueado", m, reintentar=True))
        threading.Thread(target=work, daemon=True).start()

    def _lic_ok(self, s, r=None):
        self._lic = s
        self._set_plan((r or {}).get("plan"))
        if self._lock is not None:
            try: self._lock.destroy()
            except Exception: pass
            self._lock = None
        if self._hb:
            self.after_cancel(self._hb)
        self._hb = self.after(600_000, self._lic_heartbeat)
        self.title(f"Gestiq — {s.get('email', '')}")
        self.lbl_sess.configure(text=f"👤 {s.get('email', '')}")
        self.sess_box.pack(side="right", padx=(0, 8), pady=12)

    def _lic_cerrar_sesion(self):
        """Cierra la sesión (o permite cambiar de cuenta) desde el encabezado."""
        if any(t._running for t in (self.imagine_tab, self.guardian_tab)):
            if not messagebox.askyesno(
                    "Cerrar sesión",
                    "Hay una consulta en curso. Si cierras la sesión se "
                    "detendrá.\n\n¿Cerrar sesión de todas formas?"):
                return
            for t in (self.imagine_tab, self.guardian_tab):
                if t._running:
                    t._do_stop()
        if self._hb:
            self.after_cancel(self._hb); self._hb = None
        if licencia:
            licencia.cerrar_sesion()
        self._lic = None
        self.title("Gestiq")
        self._lock_mostrar("login")

    def _lic_heartbeat(self):
        """Revalida la licencia cada 10 min. Si falla, bloquea la app."""
        if self._lic is None or licencia is None:
            return
        def work():
            try:
                r = licencia.verificar(self._lic)
                if not r.get("ok"):
                    self.after(0, lambda m=licencia.motivo(r): self._lic_bloquear(m))
                else:
                    self.after(0, lambda p=r.get("plan"): self._set_plan(p))
            except Exception as e:
                self.after(0, lambda m=str(e): self._lic_bloquear(m, reintentar=True))
        threading.Thread(target=work, daemon=True).start()
        self._hb = self.after(600_000, self._lic_heartbeat)

    def _lic_bloquear(self, msg, reintentar=False):
        """Corta cualquier proceso en curso y muestra la pantalla de bloqueo."""
        if self._hb:
            self.after_cancel(self._hb); self._hb = None
        for t in (self.imagine_tab, self.guardian_tab):
            if t._running:
                t._do_stop()
        self._lock_mostrar("bloqueado", msg, reintentar=reintentar)

    def lic_check_run(self, tab=None):
        """Validación obligatoria antes de cada ejecución del bot."""
        if licencia is None or not licencia.configurado():
            self._lock_mostrar("bloqueado",
                               "Error interno de licencias. Reinstala la aplicación "
                               "o contacta soporte.")
            return False
        if self._lic is None:
            self._lock_mostrar("login"); return False
        try:
            r = licencia.verificar(self._lic)
            if r.get("ok"):
                self._set_plan(r.get("plan"))
                mod = "guardian" if tab is self.guardian_tab else "imagine"
                if mod not in self._mods_permitidos():
                    messagebox.showinfo(
                        "Módulo no incluido",
                        "Tu plan actual no incluye este módulo.\n"
                        "Puedes ampliarlo en la página de tu cuenta.")
                    return False
                return True
            self._lic_bloquear(licencia.motivo(r))
        except Exception as e:
            self._lic_bloquear(str(e), reintentar=True)
        return False

    def _lock_mostrar(self, modo, msg="", reintentar=False):
        """Overlay a pantalla completa: cargando / login / bloqueado."""
        try: self.sess_box.pack_forget()
        except Exception: pass
        if self._lock is not None:
            try: self._lock.destroy()
            except Exception: pass
        self._lic = None if modo != "cargando" else self._lic
        self._lock = ctk.CTkFrame(self, fg_color=BG, corner_radius=0)
        self._lock.place(relx=0, rely=0, relwidth=1, relheight=1)
        self._lock.lift()

        card = ctk.CTkFrame(self._lock, fg_color=CARD, corner_radius=16,
                            border_width=1, border_color=BORDER)
        card.place(relx=.5, rely=.45, anchor="center")
        inner = ctk.CTkFrame(card, fg_color="transparent")
        inner.pack(padx=44, pady=36)

        ctk.CTkLabel(inner, text="🔒", font=F(34)).pack()

        if modo == "cargando":
            ctk.CTkLabel(inner, text="Verificando licencia…", font=F(17, True),
                         text_color=TX).pack(pady=(10, 4))
            ctk.CTkLabel(inner, text="Conectando con el servidor",
                         font=F(12), text_color=TM).pack()

        elif modo == "login":
            ctk.CTkLabel(inner, text="Inicia sesión en Gestiq", font=F(17, True),
                         text_color=TX).pack(pady=(10, 2))
            ctk.CTkLabel(inner, text="Usa la cuenta que recibiste con tu suscripción",
                         font=F(12), text_color=TM).pack(pady=(0, 14))
            self._lock_email = ctk.CTkEntry(inner, width=290, height=40,
                                            placeholder_text="Correo",
                                            font=F(13), fg_color=CARD2,
                                            border_color=BORDER, text_color=TX)
            self._lock_email.pack(pady=4)
            self._lock_email.insert(0, licencia.ultimo_email())
            self._lock_pass = ctk.CTkEntry(inner, width=290, height=40,
                                           placeholder_text="Contraseña", show="•",
                                           font=F(13), fg_color=CARD2,
                                           border_color=BORDER, text_color=TX)
            self._lock_pass.pack(pady=4)
            self._lock_pass.bind("<Return>", lambda *_: self._lic_login())
            self._lock_btn = ctk.CTkButton(inner, text="Entrar", width=290, height=42,
                                           font=F(14, True), fg_color=ACCENT,
                                           hover_color=ACCENT_H,
                                           command=self._lic_login)
            self._lock_btn.pack(pady=(12, 4))
            self._lock_status = ctk.CTkLabel(inner, text="", font=F(12),
                                             text_color=ERR, wraplength=290)
            self._lock_status.pack()
            ctk.CTkButton(inner, text="¿No tienes cuenta? Regístrate y paga aquí",
                          width=290, height=34, font=F(12),
                          fg_color="transparent", hover_color=CARD2,
                          text_color=ACCENT,
                          command=lambda: webbrowser.open(REGISTRO_URL)
                          ).pack(pady=(8, 0))
            ctk.CTkLabel(inner, text="Se requiere conexión a internet",
                         font=F(11), text_color=TM).pack(pady=(10, 0))

        else:   # bloqueado
            ctk.CTkLabel(inner, text="Acceso bloqueado", font=F(17, True),
                         text_color=TX).pack(pady=(10, 6))
            ctk.CTkLabel(inner, text=msg or "Licencia no válida.", font=F(13),
                         text_color=TM, wraplength=320, justify="center").pack()
            if reintentar:
                ctk.CTkButton(inner, text="Reintentar", width=290, height=42,
                              font=F(14, True), fg_color=ACCENT,
                              hover_color=ACCENT_H,
                              command=self._lic_iniciar).pack(pady=(16, 4))
            ctk.CTkButton(inner, text="Cambiar de cuenta", width=290, height=38,
                          font=F(13), fg_color=CARD2, hover_color=BORDER,
                          text_color=TX,
                          command=lambda: (licencia.cerrar_sesion(),
                                           self._lock_mostrar("login"))
                          ).pack(pady=(6 if not reintentar else 0, 0))

    def _lic_login(self):
        em = self._lock_email.get().strip()
        pw = self._lock_pass.get()
        if not em or not pw:
            self._lock_status.configure(text="Escribe tu correo y contraseña."); return
        self._lock_btn.configure(state="disabled", text="Verificando…")
        self._lock_status.configure(text="")
        def work():
            try:
                s = licencia.login(em, pw)
                r = licencia.verificar(s)
                if r.get("ok"):
                    self.after(0, lambda r=r: self._lic_ok(s, r))
                else:
                    licencia.cerrar_sesion()
                    self.after(0, lambda m=licencia.motivo(r):
                               self._lock_mostrar("bloqueado", m))
            except Exception as e:
                def ui(m=str(e)):
                    try:
                        self._lock_btn.configure(state="normal", text="Entrar")
                        self._lock_status.configure(text=m)
                    except Exception:
                        pass
                self.after(0, ui)
        threading.Thread(target=work, daemon=True).start()

    def log(self, tab, msg, lvl="info"):
        """Escribe en el registro de la pestaña. Thread-safe."""
        colors   = {"info": LOG_FG, "ok": OK_C, "warn": WA_C, "error": ER_C, "orange": OR_C}
        prefixes = {"info": "·  ", "ok": "✓  ", "warn": "⚠  ", "error": "✗  ", "orange": "◆  "}
        ts   = datetime.now().strftime("%H:%M:%S")
        line = f"{ts}  {prefixes.get(lvl, '   ')}{msg}\n"
        color = colors.get(lvl, LOG_FG)

        def _write():
            w = tab.log_text
            w.configure(state="normal")
            # una etiqueta fija por nivel (id() se reutiliza y mezclaba colores)
            tag = f"lvl_{lvl}"
            w.tag_config(tag, foreground=color)
            w.insert("end", line, tag)
            w.see("end")
            w.configure(state="disabled")
        tab.after(0, _write)


# ════════════════════════════════════════════════════════════════════════════
#  Tab base
# ════════════════════════════════════════════════════════════════════════════
class BaseTab(ctk.CTkFrame):
    def __init__(self, parent, app):
        super().__init__(parent, fg_color="transparent")
        self.app       = app
        self.xl_path   = None
        self.wb        = None
        self._stop     = False
        self._running  = False
        self._controls = []          # widgets que se bloquean durante la ejecución
        self._loop     = None        # event loop del hilo de automatización
        self._task     = None        # tarea asyncio en curso (para cancelarla)
        self._thread   = None        # hilo de automatización (vigilado desde la UI)
        self._login_ev  = None       # espera del login manual
        self._login_dlg = None
        self._setup_ui()

    # ── UI ──────────────────────────────────────────────────────────────────
    def _setup_ui(self):
        self.columnconfigure(0, weight=1)
        self.rowconfigure(2, weight=1)   # la tarjeta de ejecución se expande

        # ── Tarjeta 1 · Archivo ──────────────────────────────────────────────
        card = ctk.CTkFrame(self, fg_color=CARD, corner_radius=12)
        card.grid(row=0, column=0, sticky="ew", pady=(0, 14))
        card.columnconfigure(0, weight=1)

        ctk.CTkLabel(card, text="1 · Archivo Excel", font=F(13, True),
                     text_color=TX, anchor="w").grid(
            row=0, column=0, sticky="w", padx=20, pady=(16, 8))

        zone = ctk.CTkFrame(card, fg_color=CARD2, corner_radius=10)
        zone.grid(row=1, column=0, sticky="ew", padx=20)
        zone.columnconfigure(1, weight=1)

        ctk.CTkLabel(zone, text="📄", font=F(20)).grid(
            row=0, column=0, rowspan=2, padx=(16, 10), pady=14)

        self.lbl_file = ctk.CTkLabel(zone, text="Ningún archivo seleccionado",
                                     font=F(13), text_color=TM, anchor="w")
        self.lbl_file.grid(row=0, column=1, sticky="ew", pady=(12, 0))

        self.lbl_file_info = ctk.CTkLabel(zone, text="Elige el Excel que contiene los casos",
                                          font=F(11), text_color=TM, anchor="w",
                                          justify="left")
        self.lbl_file_info.grid(row=1, column=1, sticky="ew", pady=(0, 12))

        # ajustar el texto al ancho disponible en vez de cortarlo
        # (se mide la zona completa: su ancho no depende del texto, evita bucles)
        def _fit_labels(e):
            avail = max(160, e.width - 230)   # icono + botón + márgenes
            self.lbl_file.configure(wraplength=avail)
            self.lbl_file_info.configure(wraplength=avail)
        zone.bind("<Configure>", _fit_labels)

        self.btn_select = ctk.CTkButton(zone, text="Seleccionar…",
                                        command=self._pick_file,
                                        fg_color=ACCENT, hover_color=ACCENT_H,
                                        text_color="#FFFFFF", font=F(13, True),
                                        corner_radius=10, height=36, width=130)
        self.btn_select.grid(row=0, column=2, rowspan=2, padx=16)
        self._controls.append(self.btn_select)

        ctk.CTkLabel(card, text="Los resultados se escriben en este archivo; las celdas que ya tienen valor no se tocan.",
                     font=F(11), text_color=TM, anchor="w").grid(
            row=2, column=0, sticky="w", padx=20, pady=(6, 16))

        # ── Tarjeta 2 · Configuración ────────────────────────────────────────
        cfg = ctk.CTkFrame(self, fg_color=CARD, corner_radius=12)
        cfg.grid(row=1, column=0, sticky="ew", pady=(0, 14))

        ctk.CTkLabel(cfg, text="2 · Configuración", font=F(13, True),
                     text_color=TX, anchor="w").pack(fill="x", padx=20, pady=(16, 10))

        row = ctk.CTkFrame(cfg, fg_color="transparent")
        row.pack(fill="x", padx=20)

        col = self._field(row, "Hoja del Excel")
        self.v_sheet  = ctk.StringVar()
        self.cb_sheet = ctk.CTkComboBox(col, variable=self.v_sheet, width=170,
                                        height=34, state="readonly", corner_radius=8,
                                        fg_color=CARD2, border_color=BORDER,
                                        button_color=ACCENT, button_hover_color=ACCENT_H,
                                        dropdown_fg_color=CARD, dropdown_text_color=TX,
                                        dropdown_hover_color=CARD2,
                                        text_color=TX, font=F(13))
        self.cb_sheet.pack()

        self._extra_config(row)

        ctk.CTkLabel(cfg, text="ℹ  " + self._config_help(), font=F(11),
                     text_color=TM, anchor="w", justify="left",
                     wraplength=620).pack(fill="x", padx=20, pady=(8, 16))

        # ── Tarjeta 3 · Ejecución ────────────────────────────────────────────
        run = ctk.CTkFrame(self, fg_color=CARD, corner_radius=12)
        run.grid(row=2, column=0, sticky="nsew")
        run.columnconfigure(0, weight=1)
        run.rowconfigure(4, weight=1)

        ctk.CTkLabel(run, text="3 · Ejecución", font=F(13, True),
                     text_color=TX, anchor="w").grid(
            row=0, column=0, sticky="w", padx=20, pady=(16, 10))

        btns = ctk.CTkFrame(run, fg_color="transparent")
        btns.grid(row=1, column=0, sticky="ew", padx=20)

        self.btn_start = ctk.CTkButton(btns, text="▶   Iniciar consulta",
                                       command=self._start,
                                       fg_color=ACCENT, hover_color=ACCENT_H,
                                       text_color="#FFFFFF", font=F(14, True),
                                       corner_radius=10, height=44, width=200)
        self.btn_start.pack(side="left")

        self.btn_stop = ctk.CTkButton(btns, text="⏹  Detener búsqueda",
                                      command=self._do_stop,
                                      fg_color=ERR, hover_color=ERR_H,
                                      text_color="#FFFFFF", font=F(13, True),
                                      corner_radius=10, height=44, width=170,
                                      state="disabled")
        self.btn_stop.pack(side="left", padx=(12, 0))

        self.btn_download = ctk.CTkButton(btns, text="⬇  Guardar copia…",
                                          command=self._save,
                                          fg_color=CARD2, hover_color=BORDER,
                                          text_color=TX, font=F(13),
                                          corner_radius=10, height=44, width=160,
                                          state="disabled")
        self.btn_download.pack(side="right")

        prog_row = ctk.CTkFrame(run, fg_color="transparent")
        prog_row.grid(row=2, column=0, sticky="ew", padx=20, pady=(14, 0))
        prog_row.columnconfigure(0, weight=1)

        self.lbl_prog = ctk.CTkLabel(prog_row, text="Listo para comenzar",
                                     font=F(12), text_color=TM, anchor="w")
        self.lbl_prog.grid(row=0, column=0, sticky="w")

        self.lbl_pct = ctk.CTkLabel(prog_row, text="", font=F(12, True),
                                    text_color=TM, anchor="e")
        self.lbl_pct.grid(row=0, column=1, sticky="e")

        self.prog = ctk.CTkProgressBar(prog_row, height=8, fg_color=CARD2,
                                       progress_color=ACCENT, corner_radius=4)
        self.prog.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(6, 0))
        self.prog.set(0)
        self._prog_val    = 0.0     # valor mostrado (animado)
        self._prog_target = 0.0     # valor real al que se anima
        self._prog_anim   = None

        ctk.CTkLabel(run, text="Registro de actividad", font=F(11, True),
                     text_color=TM, anchor="w").grid(
            row=3, column=0, sticky="w", padx=20, pady=(14, 4))

        log_frame = ctk.CTkFrame(run, fg_color=LOG_BG, corner_radius=10)
        log_frame.grid(row=4, column=0, sticky="nsew", padx=20, pady=(0, 20))
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)

        self.log_text = tk.Text(
            log_frame,
            bg=LOG_BG, fg=LOG_FG,
            font=(MONO_FAM, 13 if sys.platform == "darwin" else 11),
            relief="flat", bd=0, highlightthickness=0,
            state="disabled", wrap="word",
            padx=12, pady=8, spacing1=3,
            insertbackground=ACCENT,
            selectbackground="#33334D",
        )
        self.log_text.grid(row=0, column=0, sticky="nsew", padx=(4, 0), pady=4)

        sb = ctk.CTkScrollbar(log_frame, command=self.log_text.yview,
                              fg_color=LOG_BG, button_color="#3B3B54",
                              button_hover_color=ACCENT)
        sb.grid(row=0, column=1, sticky="ns", pady=4, padx=(0, 4))
        self.log_text.configure(yscrollcommand=sb.set)

    # ── Helpers de UI ─────────────────────────────────────────────────────────
    def _field(self, parent, label):
        """Columna etiqueta-arriba/campo-abajo para la fila de configuración."""
        col = ctk.CTkFrame(parent, fg_color="transparent")
        col.pack(side="left", padx=(0, 14), pady=(0, 4))
        ctk.CTkLabel(col, text=label, font=F(11), text_color=TM,
                     anchor="w").pack(fill="x", pady=(0, 4))
        return col

    def _entry(self, parent, var, width):
        e = ctk.CTkEntry(parent, textvariable=var, width=width, height=34,
                         fg_color=CARD2, border_color=BORDER, border_width=1,
                         text_color=TX, font=F(13), corner_radius=8)
        self._controls.append(e)
        return e

    def _set_running(self, running):
        """Bloquea/desbloquea los controles durante la ejecución."""
        self._running = running
        for w in self._controls:
            w.configure(state="disabled" if running else "normal")
        self.cb_sheet.configure(state="disabled" if running else "readonly")

    def _extra_config(self, parent):
        pass

    def _config_help(self):
        return "Ajusta los nombres solo si tu Excel usa encabezados distintos."

    # ── Archivo ──────────────────────────────────────────────────────────────
    def _pick_file(self):
        p = filedialog.askopenfilename(
            title="Seleccionar Excel",
            filetypes=[("Excel", "*.xlsx *.xls"), ("Todos", "*.*")])
        if not p: return
        try:
            wb = openpyxl.load_workbook(p)
        except Exception as e:
            messagebox.showerror(
                "No se pudo abrir",
                "El archivo no se pudo abrir.\n"
                "Ciérralo en Excel e inténtalo de nuevo.")
            self.app.log(self, f"Error al abrir archivo: {e}", "error")
            return
        self.xl_path = p
        self.wb      = wb
        sheets       = wb.sheetnames
        self.cb_sheet.configure(values=sheets)
        self.cb_sheet.set(sheets[0])
        name = os.path.basename(p)
        if len(name) > 55:                       # nombres muy largos: elipsis al medio
            name = name[:30] + "…" + name[-22:]
        self.lbl_file.configure(text=name, text_color=TX)
        self.lbl_file_info.configure(text=f"{len(sheets)} hoja(s) disponibles")
        self.btn_select.configure(text="Cambiar")
        self.app.log(self, f"Cargado: {os.path.basename(p)}  ({len(sheets)} hoja(s))", "ok")

    # ── Inicio / parada ──────────────────────────────────────────────────────
    def _start(self):
        if not self.xl_path:
            messagebox.showwarning("Falta el archivo",
                                   "Primero selecciona un archivo Excel.")
            return
        if not self.v_sheet.get():
            messagebox.showwarning("Falta la hoja",
                                   "Selecciona la hoja del Excel que quieres procesar.")
            return
        if not HAVE_PW:
            messagebox.showerror("Falta un componente",
                                 "Playwright no está instalado en este equipo.")
            return
        if not self.app.lic_check_run(self):
            return
        try:
            self.wb = openpyxl.load_workbook(self.xl_path)
            sheets  = self.wb.sheetnames
            self.cb_sheet.configure(values=sheets)
        except Exception as e:
            messagebox.showerror(
                "No se pudo abrir",
                "El archivo no se pudo recargar.\n"
                "Ciérralo en Excel e inténtalo de nuevo.")
            self.app.log(self, f"Error al recargar archivo: {e}", "error")
            return
        self._stop = False
        self._login_ev = self._login_dlg = None
        self._set_running(True)
        self.btn_start.configure(state="disabled")
        self.btn_stop.configure(state="normal")
        self.btn_download.configure(state="disabled")
        if self._prog_anim:
            self.after_cancel(self._prog_anim)
            self._prog_anim = None
        self._prog_val = self._prog_target = 0.0
        self.prog.set(0)
        self.lbl_pct.configure(text="0 %")
        self.lbl_prog.configure(text="Iniciando…", text_color=TM)
        self._thread = threading.Thread(target=self._runner, daemon=True)
        self._thread.start()
        self.after(200, self._watch_thread)

    def _watch_thread(self):
        """Vigila el hilo de automatización desde el hilo principal y
        restaura la interfaz en cuanto termina (funciona en toda plataforma)."""
        if self._thread is not None and self._thread.is_alive():
            self.after(150, self._watch_thread)
        else:
            self._thread = None
            self._on_finish()

    def _runner(self):
        """Ejecuta la automatización en su propio event loop, cancelable al instante."""
        loop = asyncio.new_event_loop()
        self._loop = loop
        asyncio.set_event_loop(loop)
        self._task = loop.create_task(self._automate())
        try:
            loop.run_until_complete(self._task)
        except asyncio.CancelledError:
            pass                                    # detenido por el usuario
        except Exception as e:
            self.app.log(self, f"Error inesperado: {e}", "error")
        finally:
            try:
                loop.run_until_complete(loop.shutdown_asyncgens())
            except Exception:
                pass
            loop.close()
            self._loop = self._task = None
            # _watch_thread (hilo principal) detecta el fin y llama _on_finish

    def _do_stop(self):
        """Detiene la búsqueda de inmediato: aborta el caso en curso y cierra el navegador."""
        self._stop = True
        self.btn_stop.configure(state="disabled")
        self.lbl_prog.configure(text="Deteniendo…", text_color=TM)

        # Si está esperando el login manual, libera esa espera y cierra el diálogo
        dlg, ev = self._login_dlg, self._login_ev
        self._login_dlg = None
        if dlg is not None:
            try:
                if dlg.winfo_exists():
                    dlg.destroy()
            except Exception:
                pass
        if ev is not None and not ev.is_set():
            ev.set()        # _automate ve _stop=True y cierra el navegador ordenadamente
        else:
            # Cancela la tarea asyncio: aborta la petición en curso sin esperar a que termine
            loop, task = self._loop, self._task
            if loop is not None and task is not None and not task.done():
                loop.call_soon_threadsafe(task.cancel)
        self.app.log(self, "Detenido por el usuario. Lo ya consultado quedó en el Excel.", "warn")

    def _on_finish(self):
        self._set_running(False)
        self.btn_start.configure(state="normal")
        self.btn_stop.configure(state="disabled")
        self.btn_download.configure(state="normal")
        if self._stop:
            self.lbl_prog.configure(text="Detenido ⏹ — puedes guardar lo procesado hasta ahora",
                                    text_color=WA_C)
        else:
            self.lbl_prog.configure(text="Completado ✓ — revisa el registro y guarda una copia",
                                    text_color=ACCENT)
            self._set_prog(1)

    def _upd(self, cur, tot, msg=""):
        pct = cur / tot if tot else 0
        self._set_prog(pct)
        self.lbl_prog.configure(text=f"Procesando: {msg}  —  {cur} de {tot}",
                                text_color=TM)

    def _set_prog(self, target):
        """Anima la barra suavemente hacia el valor objetivo."""
        self._prog_target = max(0.0, min(1.0, target))
        if self._prog_anim is None:
            self._animate_prog()

    def _animate_prog(self):
        try:
            d = self._prog_target - self._prog_val
            if abs(d) < 0.002:
                self._prog_val  = self._prog_target
                self._prog_anim = None
            else:
                self._prog_val += d * 0.12          # easing exponencial
                self._prog_anim = self.after(16, self._animate_prog)
            self.prog.set(self._prog_val)
            self.lbl_pct.configure(text=f"{int(self._prog_val * 100)} %")
        except tk.TclError:
            self._prog_anim = None                  # ventana cerrada

    # ── Guardar Excel ────────────────────────────────────────────────────────
    def _save(self):
        if not self.wb: return
        p = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"resultado_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
        if not p: return
        try:
            self.wb.save(p)
            messagebox.showinfo("Guardado", f"Archivo guardado en:\n{p}")
            self.app.log(self, f"Excel guardado: {os.path.basename(p)}", "ok")
        except Exception as e:
            messagebox.showerror(
                "No se pudo guardar",
                "Revisa que el archivo de destino no esté abierto en Excel.")
            self.app.log(self, f"Error al guardar: {e}", "error")

    # ── Diálogo de login ─────────────────────────────────────────────────────
    def _ask_ready(self, event, system_name):
        self.lbl_prog.configure(text="Esperando login manual…", text_color=TM)

        dlg = ctk.CTkToplevel(self)
        self._login_ev, self._login_dlg = event, dlg
        dlg.title("Inicia sesión para continuar")
        dlg.configure(fg_color=BG)
        dlg.resizable(False, False)
        self.update_idletasks()
        x = self.winfo_rootx() + max((self.winfo_width() - 520) // 2, 0)
        y = self.winfo_rooty() + 100
        dlg.geometry(f"520x280+{max(x, 0)}+{max(y, 0)}")
        dlg.grab_set()
        dlg.transient(self.app)
        dlg.lift()

        ctk.CTkLabel(dlg, text="🌐", font=F(28)).pack(pady=(22, 2))
        ctk.CTkLabel(dlg, text="Se abrió el navegador", font=F(16, True),
                     text_color=TX).pack()
        ctk.CTkLabel(dlg,
            text=f"1.  Inicia sesión en {system_name}\n"
                 f"2.  Cuando estés dentro, vuelve a esta ventana\n"
                 f"3.  Pulsa «Continuar»",
            font=F(13), text_color=TM, justify="left",
            wraplength=440).pack(pady=10)

        bf = ctk.CTkFrame(dlg, fg_color="transparent")
        bf.pack(pady=(4, 20))

        def _ok():
            self._login_dlg = None
            self.lbl_prog.configure(text="Procesando…", text_color=TM)
            dlg.destroy(); event.set()
        def _cancel():
            self._login_dlg = None
            self._stop = True; dlg.destroy(); event.set()
        dlg.protocol("WM_DELETE_WINDOW", _cancel)

        ctk.CTkButton(bf, text="▶  Continuar", command=_ok,
                      fg_color=ACCENT, hover_color=ACCENT_H,
                      text_color="#FFFFFF", font=F(13, True),
                      corner_radius=10, height=40, width=160).pack(side="left", padx=8)
        ctk.CTkButton(bf, text="Cancelar", command=_cancel,
                      fg_color=CARD2, hover_color=BORDER, text_color=TX,
                      font=F(13), corner_radius=10, height=40,
                      width=110).pack(side="left", padx=8)

    # ── Helpers Excel ────────────────────────────────────────────────────────
    def _find_header_row(self, ws):
        for i, row in enumerate(ws.iter_rows(max_row=10), 1):
            if sum(1 for c in row if c.value) >= 3:
                return i
        return 1

    def _find_col(self, ws, names, hdr_row=None, last=False):
        hr    = hdr_row or self._find_header_row(ws)
        found = None
        for cell in ws[hr]:
            if cell.value and str(cell.value).strip().upper() in [n.upper() for n in names]:
                if not last: return cell.column
                found = cell.column
        return found

    def _get_or_create_col(self, ws, name, hdr_row, last=False):
        col = self._find_col(ws, [name], hdr_row, last=last)
        if col: return col
        col = ws.max_column + 1
        ws.cell(row=hdr_row, column=col, value=name)
        return col

    async def _automate(self): pass


# ════════════════════════════════════════════════════════════════════════════
#  Pestaña IMAGINE
# ════════════════════════════════════════════════════════════════════════════
IMAGINE_BASE = "https://imaginex-intranet.segurosbolivar.com"


class ImagineTab(BaseTab):

    def _extra_config(self, parent):
        col = self._field(parent, "Columna de casos")
        self.v_caso = ctk.StringVar(value="CASO")
        self._entry(col, self.v_caso, 130).pack()

        col = self._field(parent, "Columna de resultado")
        self.v_out = ctk.StringVar(value="IMAGINE")
        self._entry(col, self.v_out, 130).pack()

    def _config_help(self):
        return ("El bot consulta cada caso en IMAGINE y escribe la última "
                "observación en la columna de resultado. Ajusta los nombres solo "
                "si tu Excel usa encabezados distintos.")

    # ── Orquestador ──────────────────────────────────────────────────────────
    async def _automate(self):
        ws    = self.wb[self.v_sheet.get()]
        hdr   = self._find_header_row(ws)
        cc    = self._find_col(ws, ["CASO", self.v_caso.get().strip()], hdr)
        out_c = self._get_or_create_col(ws, self.v_out.get().strip().upper(), hdr)

        if not cc:
            self.after(0, lambda: messagebox.showerror("Error", "Columna CASO no encontrada."))
            self.after(0, self._on_finish); return

        jobs = []
        for row in ws.iter_rows(min_row=hdr + 1):
            cv = str(row[cc - 1].value).strip() if row[cc - 1].value else ""
            ov = str(ws.cell(row=row[0].row, column=out_c).value).strip().upper() \
                 if ws.cell(row=row[0].row, column=out_c).value else ""
            if cv and cv.upper() not in ("NONE", "NAN"):
                if not ov or ov == "PENDIENTE":
                    jobs.append((row[0].row, cv))

        total = len(jobs)
        if not total:
            self.app.log(self, "No hay casos pendientes para procesar.", "warn")
            self.after(0, self._on_finish); return

        self.app.log(self, f"{total} casos a procesar.", "info")
        not_found = []

        async with async_playwright() as pw:
            browser = await pw.chromium.launch(headless=False, slow_mo=250)
            ctx     = await browser.new_context()
            page    = await ctx.new_page()
            await page.goto(IMAGINE_BASE)

            ev = threading.Event()
            self.after(0, lambda: self._ask_ready(ev, "Imagine (solo inicia sesión — el bot navega solo)"))
            ev.wait()

            if self._stop:
                await browser.close(); self.after(0, self._on_finish); return

            for i, (rn, caso) in enumerate(jobs):
                if self._stop: break
                self.after(0, lambda i=i, c=caso: (
                    self._upd(i + 1, total, f"Caso {c}"),
                    self.app.log(self, f"→ Caso {c}", "info")
                ))
                try:
                    result = await self._procesar_caso(ctx, caso)
                except Exception as e:
                    if "Sesión expirada" in str(e):
                        self.app.log(self, "Sesión expirada — proceso detenido. "
                                           "Vuelve a iniciar sesión y reinicia.", "error")
                        break
                    result = f"ERROR: {str(e)[:80]}"
                    self.app.log(self, str(e), "error")

                if result is None:
                    self.app.log(self, f"  {caso}: PENDIENTE — omitido", "warn")
                    continue

                cell = ws.cell(row=rn, column=out_c)
                cell.value = result
                rl = result.lower()
                if any(x in rl for x in ["autorizado", "aprobado", "pre factura"]):
                    style_cell(cell, F_GREEN)
                elif any(x in rl for x in ["devuelve", "rechaza", "no coincide"]):
                    style_cell(cell, F_RED)
                elif any(x in rl for x in ["errado", "no encontrado", "error"]):
                    style_cell(cell, F_ORANGE)
                    if "no encontrado" in rl: not_found.append(caso)
                else:
                    style_cell(cell)

                self.app.log(self, f"  {caso}: {result[:80]}",
                             "error" if "no encontrado" in rl else "ok")

            await browser.close()

        if not_found:
            self.app.log(self, f"No encontrados: {', '.join(not_found)}", "warn")
        self.app.log(self, f"Listo. {min(i+1,total)}/{total} procesados.", "ok")
        self.after(0, self._on_finish)

    # ── Lógica de un caso (protocolo AJAX directo — verificado en el sitio) ──
    # cargaFuncion.php es solo un "cascarón" que carga todo por AJAX; el bot
    # llama directamente los 3 endpoints (POST /arp/radicador/ajax/_radicador.php):
    #   1. opcion=cargaInfoRadicacion&na=<caso>&cons=si&band=0
    #      → info del caso (No. Caso, Estado, Razón Empresa)
    #      → si no existe: "No hay resultados para mostrar..."
    #   2. opcion=cargaHistGestion&na=<caso>&adic=0&band=0
    #      → tabla de actividades table-striped [Actividad|Usuario|Estado|...]
    #   3. opcion=cargaObsGestionTotal&na=<caso>
    #      → tabla de observaciones editorTip [Fecha|Usuario|Actividad|Observacion|Adjunto]
    # La sesión viaja en las cookies del contexto de Playwright (login manual).
    @staticmethod
    def _strip(html):
        return " ".join(re.sub(r"<[^>]+>", " ", html).split())

    @staticmethod
    def _filas(html_tabla):
        """Devuelve las filas <tr> de una tabla como listas de celdas (texto plano)."""
        filas = []
        for tr in re.findall(r"<tr[^>]*>(.*?)</tr>", html_tabla, re.S | re.I):
            tds = re.findall(r"<td[^>]*>(.*?)</td>", tr, re.S | re.I)
            filas.append([ImagineTab._strip(td) for td in tds])
        return filas

    async def _ajax(self, ctx, **datos):
        r = await ctx.request.post(
            f"{IMAGINE_BASE}/arp/radicador/ajax/_radicador.php",
            form=datos, timeout=30000)
        html = await r.text()
        low = self._strip(html).lower()
        # Sesión expirada → el servidor devuelve la página de login
        if "ingrese su usuario" in low or "autenticaci" in low:
            raise RuntimeError("Sesión expirada — vuelve a iniciar sesión en el navegador")
        return html

    async def _procesar_caso(self, ctx, caso):
        caso = str(caso).split(".")[0].strip()

        # 1) Info del caso
        html = await self._ajax(ctx, opcion="cargaInfoRadicacion",
                                na=caso, cons="si", band="0")
        body = self._strip(html)

        if "No hay resultados" in body or f"No. Caso: {caso}" not in body:
            return "CASO NO ENCONTRADO"

        # Empresa: texto entre "Razón Empresa:" y "Nit/CC"
        # (\S*n tolera problemas de codificación en la tilde de "Razón")
        m = re.search(r"Raz\S*n Empresa:\s*(.*?)\s*Nit/CC", body)
        empresa = m.group(1).upper() if m else body.upper()
        if "IPRECON" not in empresa:
            return "CASO ERRADO"

        # 2) Actividades: tabla class="table table-striped..." →
        # [Actividad | Usuario | Estado | Fecha Asig. | Fecha Cierre | Gestionar]
        html_act = await self._ajax(ctx, opcion="cargaHistGestion",
                                    na=caso, adic="0", band="0")
        mt = re.search(r'<table[^>]*table-striped[^>]*>(.*?)</table>', html_act, re.S | re.I)
        if mt:
            for celdas in self._filas(mt.group(1)):
                if len(celdas) >= 3 and "pendiente" in celdas[2].lower():
                    return None                      # actividad pendiente → omitir

        # 3) Observaciones (mismo endpoint que usa el botón del modal)
        html2 = await self._ajax(ctx, opcion="cargaObsGestionTotal", na=caso)

        # Tabla editorTip: [Fecha | Usuario | Actividad | Observacion | Adjunto]
        mo = re.search(r'<table[^>]*editorTip[^>]*>(.*?)</table>', html2, re.S | re.I)
        obs_text, best_fecha = "", ""
        if mo:
            for celdas in self._filas(mo.group(1)):
                if len(celdas) < 4:
                    continue
                fecha, obs = celdas[0], celdas[3]
                # solo filas cuya 1ª celda es fecha real (salta el encabezado)
                if not re.match(r"\d{4}-\d{2}-\d{2}", fecha):
                    continue
                if obs and fecha >= best_fecha:
                    best_fecha, obs_text = fecha, obs

        if not obs_text:
            obs_text = "SIN OBSERVACION"
        return obs_text


# ════════════════════════════════════════════════════════════════════════════
#  Pestaña GUARDIAN
# ════════════════════════════════════════════════════════════════════════════
GUARDIAN_API = "https://gestion-api.guardiandelaproductividad.com"


class GuardianTab(BaseTab):

    def _extra_config(self, parent):
        col = self._field(parent, "Columna cronograma")
        self.v_cron = ctk.StringVar(value="CRONOGRAMA")
        self._entry(col, self.v_cron, 130).pack()

        col = self._field(parent, "Columna secuencia")
        self.v_sec = ctk.StringVar(value="SECUENCIA")
        self._entry(col, self.v_sec, 120).pack()

        col = self._field(parent, "Columna de resultado")
        self.v_out = ctk.StringVar(value="GUARDIAN")
        self._entry(col, self.v_out, 120).pack()

    def _config_help(self):
        return ("El bot busca cada cronograma y secuencia en Guardián y escribe "
                "el estado (APROBADO, RECHAZADO, PENDIENTE…) en la columna de "
                "resultado.")

    # ── Orquestador ──────────────────────────────────────────────────────────
    async def _automate(self):
        ws    = self.wb[self.v_sheet.get()]
        hdr   = self._find_header_row(ws)
        cc    = self._find_col(ws, ["CRONOGRAMA", self.v_cron.get().strip()], hdr)
        sc    = self._find_col(ws, ["SECUENCIA",  self.v_sec.get().strip()],  hdr)
        out_c = self._get_or_create_col(ws, self.v_out.get().strip().upper(), hdr)

        self.app.log(self, f"Encabezado fila {hdr} | CRON col {cc} | SEC col {sc} | salida col {out_c}", "info")

        if not cc or not sc:
            self.after(0, lambda: messagebox.showerror("Error", "Columnas CRONOGRAMA o SECUENCIA no encontradas."))
            self.after(0, self._on_finish); return

        jobs = []
        for row in ws.iter_rows(min_row=hdr + 1):
            cv = str(row[cc - 1].value).strip() if row[cc - 1].value else ""
            sv = str(row[sc - 1].value).strip() if row[sc - 1].value else ""
            ov = str(ws.cell(row=row[0].row, column=out_c).value).strip().upper() \
                 if ws.cell(row=row[0].row, column=out_c).value else ""
            if cv and sv and cv.upper() not in ("NONE", "NAN", ""):
                if not ov or "PENDIENTE" in ov:
                    jobs.append((row[0].row, cv, sv))

        total = len(jobs)
        if not total:
            self.app.log(self, "No hay filas pendientes para procesar.", "warn")
            self.after(0, self._on_finish); return

        self.app.log(self, f"{total} filas a procesar.", "info")
        no_resultado = []

        async with async_playwright() as pw:
            browser = await pw.chromium.launch(headless=False, slow_mo=200)
            ctx     = await browser.new_context()
            page    = await ctx.new_page()
            await page.goto("https://www.guardiandelaproductividad.com/login/signin",
                            timeout=60000, wait_until="domcontentloaded")

            ev = threading.Event()
            self.after(0, lambda: self._ask_ready(ev, "Guardián (solo inicia sesión — el bot consulta por API)"))
            ev.wait()

            if self._stop:
                await browser.close(); self.after(0, self._on_finish); return

            # Token y usuario desde las cookies de la sesión iniciada
            cookies = {c["name"]: c["value"]
                       for c in await ctx.cookies("https://www.guardiandelaproductividad.com")}
            tok = unquote(cookies.get("sesion", ""))
            uid = unquote(cookies.get("user_id", ""))
            # la cookie user_id viene en base64 (p.ej. "NDMyMDM=" → "43203")
            if uid and not uid.isdigit():
                try:    uid = base64.b64decode(uid).decode().strip()
                except Exception: pass
            if not uid.isdigit():
                self.app.log(self, f"user_id no reconocido: {uid!r}", "error")
                await browser.close(); self.after(0, self._on_finish); return
            if not tok or not uid:
                self.app.log(self, "No se encontró la sesión — ¿iniciaste sesión?", "error")
                await browser.close(); self.after(0, self._on_finish); return
            hdrs = {"Authorization": f"Bearer {tok}", "Content-Type": "application/json"}
            self.app.log(self, f"Sesión OK (usuario {uid}) — consultando por API…", "ok")

            for i, (rn, cron, sec) in enumerate(jobs):
                if self._stop: break
                self.after(0, lambda i=i, c=cron, s=sec:
                    self._upd(i + 1, total, f"Cron {c} / Sec {s}")
                )
                try:
                    result = await self._procesar_fila_api(ctx, hdrs, uid, cron, sec)
                except Exception as e:
                    if "Sesión expirada" in str(e):
                        self.app.log(self, "Sesión expirada — proceso detenido. "
                                           "Vuelve a iniciar sesión y reinicia.", "error")
                        break
                    result = f"ERROR: {str(e)[:80]}"
                    self.app.log(self, result, "error")

                cell = ws.cell(row=rn, column=out_c)
                cell.value = result
                ru = result.upper()
                style_cell(cell)
                if "RECHAZADO" in ru:
                    style_cell(cell, F_RED, white=True);    lvl = "error"
                elif "PENDIENTE" in ru:
                    style_cell(cell, F_YELLOW);             lvl = "warn"
                elif "APROBADO" in ru:
                    style_cell(cell, F_GREEN);              lvl = "ok"
                elif "NO NECESITA" in ru:
                    style_cell(cell, F_ORANGE);             lvl = "ok"
                elif "NO ENCONTR" in ru or "ACTIVIDAD NO" in ru:
                    no_resultado.append(f"{cron}/{sec}"); lvl = "error"
                else:
                    lvl = "orange"   # estado desconocido → visible, nunca gris

                self.app.log(self, f"{cron}/{sec}: {result}", lvl)

            await browser.close()

        if no_resultado:
            self.app.log(self, f"Sin resultado ({len(no_resultado)}): {', '.join(no_resultado)}", "warn")
        self.app.log(self, f"Listo. {min(i+1,total)}/{total} procesados.", "ok")
        self.after(0, self._on_finish)

    # ── Lógica de una fila (API directa — descubierta del sitio real) ───────
    # 1) GET  /plan-trabajo-anual/api/v1/cronograma-actividades-proveedor
    #         ?scheduleNumber&sequenceNumber&userId&size&page&sortDirection
    #    → data.data[0] = {id, scheduleId, ...}
    # 2) POST /comunes/api/v1/consultaranexos
    #         {type, parentModelId=scheduleId, modelId=id, createdBy=userId, ...}
    #    → body.rows[] = {approvalStatus:{value,title}, createdAt}
    async def _procesar_fila_api(self, ctx, hdrs, uid, cronograma, secuencia):
        cron = str(cronograma).split(".")[0].strip()
        sec  = str(secuencia).split(".")[0].strip()

        r = await ctx.request.get(
            f"{GUARDIAN_API}/plan-trabajo-anual/api/v1/cronograma-actividades-proveedor",
            params={"scheduleNumber": cron, "sequenceNumber": sec, "userId": uid,
                    "size": 5, "page": 0, "sortDirection": "asc"},
            headers=hdrs, timeout=30000)
        if r.status in (401, 403):
            raise RuntimeError("Sesión expirada")
        j = await r.json()
        acts = ((j.get("data") or {}).get("data")) or []
        if not acts:
            return "ACTIVIDAD NO ENCONTRADA"
        act = acts[0]

        r2 = await ctx.request.post(
            f"{GUARDIAN_API}/comunes/api/v1/consultaranexos",
            data=json.dumps({"type": "plan_trabajo_anual",
                             "parentModelId": str(act["scheduleId"]),
                             "modelId": str(act["id"]),
                             "size": 50, "page": 0,
                             "createdBy": int(uid),
                             "sortField": "approvalStatus",
                             "sortDirection": "",
                             "isProvider": True}),
            headers=hdrs, timeout=30000)
        if r2.status in (401, 403):
            raise RuntimeError("Sesión expirada")
        j2 = await r2.json()
        rows = ((j2.get("body") or {}).get("rows")) or []
        if not rows:
            return "NO NECESITA APROBACION"

        # documento más reciente por fecha de carga
        rows.sort(key=lambda d: str(d.get("createdAt") or ""), reverse=True)
        title = str(((rows[0].get("approvalStatus") or {}).get("title")) or "").strip()
        if not title:
            return "NO NECESITA APROBACION"
        return title.upper()  # texto tal cual lo devuelve Guardián, en mayúsculas

    # ── Lógica antigua por interfaz (ya no se usa; conservada por referencia) ─
    async def _procesar_fila(self, page, cronograma, secuencia):
        await page.wait_for_load_state("domcontentloaded")
        try:
            await page.wait_for_selector("div.spinner-container", state="hidden", timeout=10000)
        except Exception: pass
        await page.wait_for_timeout(400)

        await self._cerrar_modal(page)

        rst = await page.query_selector("button:has-text('Restablecer'), a:has-text('Restablecer')")
        if rst:
            await rst.click(); await page.wait_for_timeout(600)

        cron_inp = (
            await page.query_selector("input[name='numberSchedule']")
            or await page.query_selector("input[placeholder*='289736']")
            or await page.query_selector("input[placeholder*='cronograma' i]")
        )
        if not cron_inp:
            inputs = await page.query_selector_all("input[inputmode='decimal'], input[type='number']")
            if inputs: cron_inp = inputs[0]
        if not cron_inp:
            inputs = await page.query_selector_all("input")
            if inputs: cron_inp = inputs[0]
        if not cron_inp: return "ERROR: campo cronograma no encontrado"

        await cron_inp.click(); await page.wait_for_timeout(200)
        await cron_inp.click(click_count=3)
        await cron_inp.fill(str(cronograma))

        sec_inp = (
            await page.query_selector("input[name='numberActivity']")
            or await page.query_selector("input[placeholder='Ejemplo: 10']")
            or await page.query_selector("input[name*='activity' i]")
        )
        if not sec_inp:
            inputs = await page.query_selector_all("input[inputmode='decimal'], input[type='number']")
            if len(inputs) >= 2: sec_inp = inputs[1]
        if not sec_inp:
            inputs = await page.query_selector_all("input")
            if len(inputs) >= 2: sec_inp = inputs[1]
        if sec_inp:
            await sec_inp.click(); await page.wait_for_timeout(200)
            await sec_inp.click(click_count=3)
            await sec_inp.fill(str(secuencia))

        consultar = await page.query_selector("button:has-text('Consultar')")
        if not consultar: return "ERROR: botón Consultar no encontrado"
        await consultar.click()

        try:
            await page.wait_for_selector("div.spinner-container", state="hidden", timeout=15000)
        except Exception: pass
        await page.wait_for_timeout(800)

        body = await page.inner_text("body")
        if ("No se encontraron actividades relacionadas" in body
                or "Mostrando 1 a 0" in body
                or "no hay" in body.lower()
                or "sin resultado" in body.lower()):
            return "ACTIVIDAD NO ENCONTRADA"

        clicked = await page.evaluate("""
            () => {
                const rows = document.querySelectorAll('table tbody tr');
                if (!rows.length) return false;
                const btn = rows[0].querySelector('button, a, [role="button"], .p-button');
                if (btn) { btn.click(); return true; }
                return false;
            }
        """)

        if not clicked:
            clip = None
            for sel in ["table tbody tr:first-child td:last-child button",
                        "table tbody tr:first-child button",
                        "table tbody tr:first-child a"]:
                clip = await page.query_selector(sel)
                if clip: break
            if clip:
                try: await clip.click(timeout=5000)
                except Exception: await page.evaluate("el => el.click()", clip)
            else:
                return "ERROR: ícono de acciones (clip) no encontrado"

        try:
            await page.wait_for_selector("div.spinner-container", state="hidden", timeout=8000)
        except Exception: pass
        await page.wait_for_timeout(2000)

        modal_text = ""
        for sel in ["[role='dialog']", ".modal-content", ".modal", "dialog"]:
            m = await page.query_selector(sel)
            if m: modal_text = await m.inner_text(); break
        if not modal_text: modal_text = await page.inner_text("body")

        if ("no hay documentos" in modal_text.lower()
                or "Mostrando 1 a 0" in modal_text
                or "sin documentos" in modal_text.lower()):
            await self._cerrar_modal(page)
            return "NO NECESITA APROBACION"

        status_raw = ""
        for sel in ["[role='dialog'] tbody tr", ".modal tbody tr", "dialog tbody tr"]:
            trs = await page.query_selector_all(sel)
            if trs:
                tds = await trs[0].query_selector_all("td")
                for ci in [2, 1, 3]:
                    if ci < len(tds):
                        t = (await tds[ci].inner_text()).strip()
                        if t and len(t) > 2: status_raw = t; break
                if status_raw: break

        if not status_raw:
            mu = modal_text.upper()
            if "APROBADO" in mu:          status_raw = "Aprobado"
            elif "RECHAZADO" in mu:       status_raw = "Rechazado"
            elif "PENDIENTE APROBACI" in mu: status_raw = "Pendiente aprobación"
            else:                         status_raw = "No necesita aprobación"

        su = status_raw.upper()
        if "APROBADO" in su and "PENDIENTE" not in su: mapped = "APROBADO"
        elif "RECHAZADO" in su:                         mapped = "RECHAZADO"
        elif "PENDIENTE" in su:                         mapped = "PENDIENTE"
        else:                                           mapped = "NO NECESITA APROBACION"

        await self._cerrar_modal(page)
        return mapped

    async def _cerrar_modal(self, page):
        await page.evaluate("""
            const btns = document.querySelectorAll(
                'button.p-dialog-header-close, button[class*="p-dialog-header-close"]'
            );
            btns.forEach(b => b.click());
        """)
        await page.wait_for_timeout(400)
        await page.keyboard.press("Escape")
        await page.wait_for_timeout(400)


# ════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    App().mainloop()
