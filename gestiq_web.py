#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Gestiq — interfaz liquid glass (pywebview).
La lógica del bot, Excel y licencias vive en bots.py (sin Tk);
este archivo es solo la capa visual (HTML/CSS real con blur).
Requiere: pip install pywebview  (además de lo de siempre)
"""

import os, sys, json, queue, threading, time, webbrowser
from datetime import datetime

import webview                    # pywebview
import openpyxl

import bots as G                  # ← toda la lógica del bot (sin Tk)
licencia = G.licencia
from version import VERSION
import updater


# ── Sustitutos sin Tk ────────────────────────────────────────────────────────
class FakeVar:
    """Imita ctk.StringVar (.get/.set)."""
    def __init__(self, v=""): self.v = v
    def get(self): return self.v
    def set(self, v): self.v = v


class MsgShim:
    """Sustituye tkinter.messagebox dentro de gestiq → toasts en la web."""
    def __init__(self, api): self.api = api
    def showinfo(self, t, m, **k):    self.api.js_toast("ok", t, m)
    def showwarning(self, t, m, **k): self.api.js_toast("warn", t, m)
    def showerror(self, t, m, **k):   self.api.js_toast("err", t, m)
    def askyesno(self, t, m, **k):    return True


# ── Pestañas "headless": bots de bots.py + puente a la UI web ────────────────
class TabWeb:
    """Mixin de UI web sobre BaseBot. El MRO hace que estos métodos tapen a
    los hooks de BaseBot; _automate, _runner y los helpers se heredan."""

    def init_comun(self, api, key):
        self.api = api
        self.key = key
        self.app = api                      # api.log(tab, msg, lvl) como App.log
        self.xl_path = None
        self.wb = None
        self._stop = False
        self._running = False
        self._loop = self._task = self._thread = None
        self._login_ev = None
        self.v_sheet = FakeVar()

    # el bot llama self.after(0, fn) desde su hilo: aquí no hay loop
    # de Tk, así que se ejecuta directo (la UI se toca vía api.js(),
    # que encola sin bloquear; ver Api._js_worker).
    def after(self, _ms, fn=None, *a):
        if fn is None: return
        try: fn(*a)
        except Exception as e:
            self.api.log(self, f"Error de interfaz: {e}", "error")

    def after_cancel(self, *_): pass

    def _set_running(self, r): self._running = r

    def _upd(self, cur, tot, msg=""):
        self.api.js(f"G.progreso({json.dumps(self.key)},{int(cur)},{int(tot)},{json.dumps(str(msg))})")

    def _set_prog(self, v):
        self.api.js(f"G.barra({json.dumps(self.key)},{float(v)})")

    def _on_finish(self):
        self._set_running(False)
        self.api.js(f"G.fin({json.dumps(self.key)},{json.dumps(bool(self._stop))})")
        if getattr(self, "resumen", None):
            self.api.js(f"G.resumen({json.dumps(self.key)},{json.dumps(self.resumen)})")

    def _ask_ready(self, event, system_name):
        self._login_ev = event
        self.api.js(f"G.pedirLogin({json.dumps(self.key)},{json.dumps(system_name)})")


class ImagineWeb(TabWeb, G.ImagineBot):
    def __init__(self, api):
        self.init_comun(api, "imagine")
        self.v_caso = FakeVar("CASO")
        self.v_out  = FakeVar("IMAGINE")


class GuardianWeb(TabWeb, G.GuardianBot):
    def __init__(self, api):
        self.init_comun(api, "guardian")
        self.v_cron = FakeVar("CRONOGRAMA")
        self.v_sec  = FakeVar("SECUENCIA")
        self.v_out  = FakeVar("GUARDIAN")


# ── API expuesta a JavaScript ────────────────────────────────────────────────
class Api:
    def __init__(self):
        self._win = None
        self._closing = False               # cerrando: no tocar más la UI
        self._lic = None
        self._pend2fa = None
        self._cid2fa = ""
        self._enrol = None
        self._enrolCid = ""
        self._plan = "completo"
        self._hb_on = False
        self._tabs = {"imagine": ImagineWeb(self), "guardian": GuardianWeb(self)}
        G.messagebox = MsgShim(self)        # los messagebox del bot → toasts
        self._js_q = queue.Queue()
        threading.Thread(target=self._js_worker, daemon=True).start()

    # ── Puente Python → JS ──
    # evaluate_js es BLOQUEANTE (espera la respuesta del WebView). Llamado en
    # mal momento —p. ej. desde el cierre de la ventana en macOS (el log de
    # _do_stop) o con la página aún sin cargar— se queda esperando para
    # siempre y la app parece pegada. Por eso NADIE llama evaluate_js directo:
    # js() encola y un único hilo de fondo (daemon) despacha; si ese hilo se
    # queda esperando, no arrastra a nadie y muere con el proceso.
    def _js_worker(self):
        while True:
            code = self._js_q.get()
            if self._closing:
                continue
            try:
                if self._win: self._win.evaluate_js(code)
            except Exception:
                pass

    def js(self, code):
        if not self._closing:
            self._js_q.put(code)

    def js_toast(self, tipo, titulo, msg):
        self.js(f"G.toast({json.dumps(tipo)},{json.dumps(str(titulo))},{json.dumps(str(msg))})")

    def log(self, tab, msg, lvl="info"):
        ts = datetime.now().strftime("%H:%M:%S")
        self.js(f"G.log({json.dumps(tab.key)},{json.dumps(ts)},{json.dumps(str(msg))},{json.dumps(lvl)})")

    # ── Licencia ──
    def _lic_dict(self, modo, **k):
        d = {"modo": modo}
        d.update(k)
        return d

    def estado_inicial(self):
        if licencia is None or not licencia.configurado():
            return self._lic_dict("bloqueado", msg="Error interno de licencias. "
                                  "Reinstala la aplicación o contacta soporte.")
        err = ""
        for intento in (1, 2):
            # La 1ª petición puede pillar la red "fría" justo al arrancar
            # (DNS/TLS aún dormidos): un fallo transitorio ya no muestra
            # "Acceso bloqueado"; se reintenta solo una vez.
            try:
                s = licencia.restaurar_sesion()
                if s is None:
                    em = licencia.ultimo_email()
                    return self._lic_dict("login", email=em,
                                          prefs=G._prefs_get(em))
                r = licencia.verificar(s)
                if r.get("ok"):
                    return self._lic_ok(s, r)
                return self._lic_dict("bloqueado", msg=licencia.motivo(r))
            except Exception as e:
                err = str(e)
                if intento == 1:
                    time.sleep(1.2)
        return self._lic_dict("bloqueado", msg=err, reintentar=True)

    def _tras_sesion(self, s, em=""):
        r = licencia.verificar(s)
        if r.get("ok"):
            return self._lic_ok(s, r)
        licencia.cerrar_sesion()
        return self._lic_dict("bloqueado", msg=licencia.motivo(r))

    def login(self, em, pw):
        try:
            s = licencia.login(em, pw)
            if s.get("requiere_2fa"):
                self._pend2fa = s
                return self._lic_dict("2fa", email=s.get("email", ""))
            return self._tras_sesion(s, em)
        except Exception as e:
            return self._lic_dict("login", msg=str(e), email=em)

    def reto_2fa(self):
        """Prepara el reto del 2do factor; en SMS, envía el código."""
        try:
            s = self._pend2fa or {}
            self._cid2fa = licencia.challenge_2fa(s.get("access_token", ""), s.get("factor_id", ""))
            return {"ok": True, "tipo": s.get("factor_tipo", "totp")}
        except Exception as e:
            return {"ok": False, "msg": str(e)}

    def verificar_2fa(self, code):
        try:
            s = self._pend2fa or {}
            ses = licencia.verify_2fa(s.get("access_token", ""), s.get("factor_id", ""),
                                      self._cid2fa, code)
            self._pend2fa = None; self._cid2fa = ""
            return self._tras_sesion(ses)
        except Exception as e:
            return self._lic_dict("2fa", msg=str(e),
                                  email=(self._pend2fa or {}).get("email", ""),
                                  tipo=(self._pend2fa or {}).get("factor_tipo", "totp"))

    def login_google(self):
        try:
            return self._tras_sesion(licencia.login_google())
        except Exception as e:
            return self._lic_dict("login", msg=str(e))

    def activar_2fa(self, tipo="totp", phone=""):
        try:
            d = licencia.enrolar_2fa(self._lic, tipo, phone)
            self._enrol = {"factor_id": d["factor_id"], "tipo": tipo}
            self._enrolCid = ""
            return {"ok": True, **d}
        except Exception as e:
            return {"ok": False, "msg": str(e)}

    def enviar_codigo_2fa(self):
        """Para SMS en alta: envía el código al teléfono del factor recién creado."""
        try:
            self._enrolCid = licencia.challenge_2fa(self._lic.get("access_token", ""),
                                                    (self._enrol or {}).get("factor_id", ""))
            return {"ok": True}
        except Exception as e:
            return {"ok": False, "msg": str(e)}

    def confirmar_2fa(self, code):
        try:
            fid = (self._enrol or {}).get("factor_id", "")
            cid = self._enrolCid or licencia.challenge_2fa(self._lic.get("access_token", ""), fid)
            self._lic = licencia.verify_2fa(self._lic.get("access_token", ""), fid, cid, code)
            self._enrol = None; self._enrolCid = ""
            return {"ok": True}
        except Exception as e:
            return {"ok": False, "msg": str(e)}

    def desactivar_2fa(self, factor_id):
        try:
            licencia.desactivar_2fa(self._lic, factor_id)
            return {"ok": True}
        except Exception as e:
            return {"ok": False, "msg": str(e)}

    def _lic_ok(self, s, r):
        self._lic = s
        self._plan = str((r or {}).get("plan") or "completo").lower()
        if self._plan not in ("imagine", "guardian", "completo"):
            self._plan = "completo"
        self._arrancar_heartbeat()
        return self._lic_dict("ok", email=s.get("email", ""), plan=self._plan,
                              prefs=self._perfil())

    def salir(self):
        for t in self._tabs.values():
            if t._running: t._do_stop()
        self._lic = None
        if licencia:
            try: licencia.cerrar_sesion()
            except Exception: pass
        em = licencia.ultimo_email() if licencia else ""
        return self._lic_dict("login", email=em, prefs=G._prefs_get(em))

    def _arrancar_heartbeat(self):
        if self._hb_on: return
        self._hb_on = True
        def bucle():
            fallos_red = 0
            while True:
                time.sleep(600)
                if self._lic is None:
                    fallos_red = 0
                    continue
                try:
                    r = licencia.verificar(self._lic)
                except Exception as e:
                    # Fallo de red transitorio: no expulsar al usuario ni matar
                    # una corrida por un blip de internet. Solo bloquea tras 3
                    # fallos seguidos (~30 min sin poder validar).
                    fallos_red += 1
                    if fallos_red >= 3:
                        fallos_red = 0
                        self._bloquear(str(e), True)
                    continue
                fallos_red = 0
                if r.get("ok"):
                    self._plan = str(r.get("plan") or self._plan).lower()
                    self.js(f"G.plan({json.dumps(self._plan)})")
                else:
                    # El servidor SÍ respondió y dijo que no → bloqueo inmediato
                    self._bloquear(licencia.motivo(r))
        threading.Thread(target=bucle, daemon=True).start()

    def _bloquear(self, msg, reintentar=False):
        for t in self._tabs.values():
            if t._running: t._do_stop()
        self._lic = None
        self.js(f"G.lic({json.dumps(self._lic_dict('bloqueado', msg=msg, reintentar=reintentar))})")

    # ── Archivo ──
    def elegir_archivo(self, m):
        tab = self._tabs[m]
        sel = self._win.create_file_dialog(
            webview.OPEN_DIALOG, file_types=("Excel (*.xlsx)", "Todos (*.*)"))  # .xls no: openpyxl no lo lee
        if not sel: return None
        p = sel[0] if isinstance(sel, (list, tuple)) else sel
        try:
            wb = openpyxl.load_workbook(p)
        except Exception as e:
            self.log(tab, f"Error al abrir archivo: {e}", "error")
            return {"error": "El archivo no se pudo abrir. Ciérralo en Excel e inténtalo de nuevo."}
        tab.xl_path, tab.wb = p, wb
        nombre = os.path.basename(p)
        self.log(tab, f"Cargado: {nombre}  ({len(wb.sheetnames)} hoja(s))", "ok")
        return {"nombre": nombre, "hojas": wb.sheetnames}

    def guardar_copia(self, m):
        tab = self._tabs[m]
        if not tab.wb: return {"error": "No hay resultados que guardar."}
        sel = self._win.create_file_dialog(
            webview.SAVE_DIALOG, file_types=("Excel (*.xlsx)",),
            save_filename=f"resultado_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
        if not sel: return None
        p = sel[0] if isinstance(sel, (list, tuple)) else sel
        try:
            self.log(tab, "Guardando… (limpiando formato del libro)", "info")
            # Al guardar: (1) mostrar la cuadrícula —la plantilla la trae oculta, por
            # eso las celdas vacías se veían blancas— y (2) quitar el relleno blanco de
            # todo lo vacío, para que la zona vacía se vea con la rejilla gris en vez
            # de bloques blancos "infinitos". El blanco puede venir de 3 sitios: celdas
            # sueltas, FILAS enteras y COLUMNAS enteras con formato (estos dos últimos
            # pintan hasta el infinito y no aparecen en iter_rows, por eso antes seguía
            # saliendo blanco). NO se oculta ni se recorta nada (datos y colores quedan
            # igual). Va inline aquí (no en gestiq.py) porque este script es el
            # __main__ y no usa caché .pyc: nunca corre una versión vieja.
            from openpyxl.styles import PatternFill
            _sin_relleno = PatternFill(fill_type=None)

            def _fill_blanco(fill):
                # ¿Es un relleno sólido blanco? (rgb, tema 0 o índice de paleta blanco)
                if not fill or fill.patternType != "solid":
                    return False
                fg = fill.fgColor
                rgb = getattr(fg, "rgb", None)
                if isinstance(rgb, str) and rgb[-6:].upper() == "FFFFFF":
                    return True
                if getattr(fg, "theme", None) == 0:
                    return True
                if getattr(fg, "indexed", None) in (1, 9):
                    return True
                return False

            from openpyxl.styles import Border
            _sin_borde = Border()

            def _tiene_borde(c):
                b = c.border
                if not b: return False
                return any(getattr(getattr(b, lado, None), "style", None)
                           for lado in ("left", "right", "top", "bottom"))

            for ws in tab.wb.worksheets:
                ws.sheet_view.showGridLines = True
                # Última columna con datos reales (límite derecho de la tabla)
                ult_col = 0
                for fila in ws.iter_rows():
                    for c in fila:
                        if c.value not in (None, "") and str(c.value).strip() != "":
                            ult_col = max(ult_col, c.column)
                # Límites de merges precalculados (comparación numérica: mucho
                # más rápido que buscar coordenadas texto por cada celda)
                mrg = [(r.min_row, r.min_col, r.max_row, r.max_col)
                       for r in ws.merged_cells.ranges]
                # 1) Celdas vacías (o solo espacios) con relleno blanco
                #    + bordes sueltos fuera de la tabla (a la derecha), p. ej. una
                #    celda con bordes olvidada en la plantilla. No se tocan bordes
                #    dentro de la tabla, ni merges, ni celdas con relleno de color.
                for fila in ws.iter_rows():
                    for celda in fila:
                        v = celda.value
                        vacia = v is None or (isinstance(v, str) and v.strip() == "")
                        if not vacia:
                            continue
                        if _fill_blanco(celda.fill):
                            celda.fill = _sin_relleno
                        if celda.column > ult_col and _tiene_borde(celda):
                            con_color = (celda.fill and celda.fill.patternType == "solid"
                                         and not _fill_blanco(celda.fill))
                            en_merge = any(a <= celda.row <= c and b <= celda.column <= d
                                           for a, b, c, d in mrg)
                            if not con_color and not en_merge:
                                celda.border = _sin_borde
                # 2) Formato de FILA/COLUMNA entera con relleno blanco (el "infinito")
                for dim in list(ws.row_dimensions.values()) + list(ws.column_dimensions.values()):
                    try:
                        if _fill_blanco(dim.fill):
                            dim.fill = _sin_relleno
                    except Exception:
                        pass
            tab.wb.save(p)
            self.log(tab, f"Excel guardado: {os.path.basename(p)}", "ok")
            return {"ruta": p}
        except Exception as e:
            self.log(tab, f"Error al guardar: {e}", "error")
            return {"error": "Revisa que el archivo de destino no esté abierto en Excel."}

    # ── Ejecución ──
    def iniciar(self, m, cfg):
        tab = self._tabs[m]
        if tab._running: return {"error": "Ya hay una consulta en curso."}
        if not tab.xl_path: return {"error": "Primero selecciona un archivo Excel."}
        # OJO: NO hacer strip() al nombre — hay hojas reales con espacio al final
        # (p.ej. "SIPAB 10 NOV ") y el nombre debe coincidir EXACTO.
        hojas = [str(h) for h in ((cfg or {}).get("hojas") or []) if str(h).strip()]
        if not hojas and (cfg or {}).get("hoja"):
            hojas = [str(cfg["hoja"])]
        if not hojas: return {"error": "Selecciona al menos una hoja del Excel."}
        if not G.HAVE_PW: return {"error": "Playwright no está instalado en este equipo."}

        # Licencia obligatoria antes de cada ejecución (igual que lic_check_run)
        if licencia is None or not licencia.configurado():
            return {"error": "Error interno de licencias."}
        if self._lic is None:
            self.js("G.lic({\"modo\":\"login\"})")
            return {"error": "Inicia sesión para continuar."}
        try:
            r = licencia.verificar(self._lic)
            if not r.get("ok"):
                self._bloquear(licencia.motivo(r))
                return {"error": "Licencia no válida."}
            self._plan = str(r.get("plan") or self._plan).lower()
            self.js(f"G.plan({json.dumps(self._plan)})")
            if self._plan != "completo" and self._plan != m:
                return {"error": "Tu plan actual no incluye este módulo."}
        except Exception as e:
            return {"error": f"No se pudo validar la licencia: {e}"}

        try:
            tab.wb = openpyxl.load_workbook(tab.xl_path)
        except Exception as e:
            self.log(tab, f"Error al recargar archivo: {e}", "error")
            return {"error": "El archivo no se pudo recargar. Ciérralo en Excel e inténtalo de nuevo."}

        tab.hojas = hojas
        tab.v_sheet.set(hojas[0])
        for k, v in (cfg or {}).items():
            if k in ("hoja", "hojas") or not str(v).strip(): continue
            var = getattr(tab, "v_" + k, None)
            if var: var.set(str(v).strip())

        # Preferencias locales por cuenta fijadas por corrida.
        _prefs = G._prefs_get(self._email())
        tab.autosave_on = bool(_prefs.get("autosave", True))
        tab.estilos = _prefs.get("estilos") or None   # personalización de resultados
        if not tab.autosave_on:
            self.log(tab, "Autoguardado desactivado (Preferencias): si cierras o "
                          "se corta, lo consultado no queda respaldado.", "warn")

        tab._stop = False
        tab._login_ev = None
        tab._set_running(True)
        tab._thread = threading.Thread(target=tab._runner, daemon=True)
        tab._thread.start()

        def vigilar(t=tab):
            t._thread.join()
            t._thread = None
            t._on_finish()
        threading.Thread(target=vigilar, daemon=True).start()
        return {"ok": True}

    def detener(self, m):
        tab = self._tabs[m]
        if tab._running: tab._do_stop()
        return {"ok": True}

    def continuar_login(self, m):
        ev = self._tabs[m]._login_ev
        if ev and not ev.is_set(): ev.set()
        return {"ok": True}

    def cancelar_login(self, m):
        tab = self._tabs[m]
        tab._stop = True
        ev = tab._login_ev
        if ev and not ev.is_set(): ev.set()
        return {"ok": True}

    # ── Perfil y preferencias (asociadas a la cuenta) ──
    def _email(self):
        return (self._lic or {}).get("email", "")

    def _perfil(self):
        """Perfil de la cuenta + preferencias de equipo (recordar cuenta)."""
        p = dict(self._perfil_cuenta())
        try:
            if licencia is not None:
                p["recordar"] = licencia.recordar_get()
        except Exception:
            pass
        return p

    def _perfil_cuenta(self):
        """Perfil de la cuenta: el servidor manda, con respaldo local.
        La primera vez tras actualizar, sube el perfil local al servidor."""
        em = self._email()
        local = G._prefs_get(em)
        if self._lic is None or licencia is None:
            return local
        try:
            serv = licencia.get_profile(self._lic)
        except Exception:
            serv = {}
        if serv:
            if any(local.get(k) != v for k, v in serv.items()):
                try: G._prefs_set(em, **serv)
                except Exception: pass
            # Claves solo-locales (p.ej. autosave) se conservan: el servidor
            # únicamente conoce nombre/foto/tema/módulo.
            r = dict(local); r.update(serv)
            return r
        if local:                      # servidor vacío → migrar lo local
            try: licencia.set_profile(self._lic, **local)
            except Exception: pass
        return local

    def prefs_get(self):
        return self._perfil()

    def prefs_set(self, kw):
        em = self._email()
        if not em:
            return {"error": "No hay sesión activa."}
        try:
            kw = dict(kw or {})
            if "recordar" in kw:            # sesión de IMAGINE/GUARDIÁN (por equipo)
                rec = bool(kw.pop("recordar"))
                try:
                    if licencia is not None and rec != licencia.recordar_get():
                        # Cambió → se aplica y se OLVIDAN las sesiones guardadas:
                        # la próxima consulta pedirá iniciar sesión de nuevo.
                        licencia.recordar_set(rec)
                        G.nav_borrar_todo()
                        self.js_toast("ok", "Inicio de sesión",
                                      "Sesión de IMAGINE/GUARDIÁN olvidada — la "
                                      "próxima consulta pedirá iniciar sesión.")
                except Exception:
                    pass
            kw = {k: v for k, v in kw.items()
                  if k in ("nombre", "foto", "tema", "modulo", "autosave", "estilos")}
            G._prefs_set(em, **kw)
            if self._lic is not None and licencia is not None:
                try: licencia.set_profile(self._lic, **kw)
                except Exception: pass
            return {"ok": True, "prefs": self._perfil()}
        except Exception as e:
            return {"error": str(e)}

    def leer_foto(self):
        """Diálogo de imagen; devuelve base64 sin procesar (recorte en JS)."""
        sel = self._win.create_file_dialog(
            webview.OPEN_DIALOG,
            file_types=("Imágenes (*.png;*.jpg;*.jpeg;*.webp;*.gif;*.bmp)",))
        if not sel:
            return None
        p = sel[0] if isinstance(sel, (list, tuple)) else sel
        try:
            if os.path.getsize(p) > 12_000_000:
                return {"error": "La imagen pesa demasiado (máx. 12 MB)."}
            import base64
            with open(p, "rb") as f:
                raw = base64.b64encode(f.read()).decode()
            ext = os.path.splitext(p)[1].lower().lstrip(".") or "png"
            if ext == "jpg":
                ext = "jpeg"
            return {"b64": raw, "mime": f"image/{ext}"}
        except Exception as e:
            return {"error": str(e)}

    # ── Varios ──
    def abrir_web(self):
        webbrowser.open(G.REGISTRO_URL)
        return {"ok": True}

    def version(self):
        return VERSION

    def exportar_log(self, m, texto):
        """Guarda el registro de actividad del módulo como .txt."""
        sel = self._win.create_file_dialog(
            webview.SAVE_DIALOG, file_types=("Texto (*.txt)",),
            save_filename=f"registro_{m}_{datetime.now().strftime('%Y%m%d_%H%M')}.txt")
        if not sel: return None
        p = sel[0] if isinstance(sel, (list, tuple)) else sel
        try:
            with open(p, "w", encoding="utf-8") as f:
                f.write(str(texto or ""))
            return {"ruta": p}
        except Exception as e:
            return {"error": str(e)}

    def buscar_update(self):
        """Comprobación manual de actualizaciones (clic en la versión)."""
        try:
            return updater.comprobar_ahora()
        except Exception as e:
            return {"error": f"No se pudo comprobar: {e}"}

    def update_listo(self):
        """¿La descarga del update ya terminó? (compat; el modal usa update_estado)."""
        try:
            return {"listo": updater.hay_update_listo()}
        except Exception:
            return {"listo": False}

    def update_estado(self):
        """Progreso real de la descarga del update (el modal lo sondea)."""
        try:
            return updater.estado_update()
        except Exception:
            return {"fase": "nada", "listo": False}

    def actualizar_ahora(self):
        """Botón «Actualizar ahora»: lanza el ayudante de instalación y cierra
        la app (el ayudante espera el PID, reemplaza y relanza). El candado
        `aplicado` del updater evita el doble lanzamiento cuando el cierre
        dispara también aplicar_y_reiniciar()."""
        try:
            if not updater.hay_update_listo():
                return {"error": "La actualización aún no termina de descargarse."}
        except Exception:
            return {"error": "No se pudo comprobar la descarga."}
        if [t for t in self._tabs.values() if t._running]:
            return {"error": "Hay una consulta en curso — deténla o espera a que termine."}
        self._closing = True
        try:
            if not updater.aplicar_y_reiniciar():
                self._closing = False
                return {"error": "No se pudo iniciar la instalación. Cierra la app normalmente."}
        except Exception as e:
            self._closing = False
            return {"error": f"No se pudo iniciar la instalación: {str(e)[:60]}"}

        def _fin():
            time.sleep(0.5)                 # deja llegar la respuesta al JS
            try: self._win.destroy()
            except Exception: pass
            time.sleep(2)
            _rematar(self)                  # por si start() no retorna (Windows)
        threading.Thread(target=_fin, daemon=True).start()
        return {"ok": True}


# ── Arranque ─────────────────────────────────────────────────────────────────
def _ruta(nombre):
    base = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base, nombre)


def _limpiar_cache_web():
    """WKWebView (macOS) cachea la página y puede servir una UI VIEJA tras
    actualizar (pasó con los logos y con la UI v1.0.17). Se limpia SOLO cuando
    cambia la versión (ahí sí hay HTML nuevo que cargar); en los arranques
    normales NO se toca nada, porque borrar la caché en cada inicio obliga a
    WKWebView a reconstruirla y hace que la app tarde en abrir. Nunca toca
    ~/Library/WebKit para conservar el localStorage (tema elegido)."""
    if sys.platform != "darwin":
        return
    marca = os.path.expanduser("~/.gestiq_uiver")
    try:
        if os.path.exists(marca) and open(marca).read().strip() == str(VERSION):
            return                              # misma versión → no limpiar (arranque rápido)
    except Exception:
        pass
    import shutil
    for bid in ("com.gestiq.app", "org.python.python", "Python"):
        shutil.rmtree(os.path.expanduser(f"~/Library/Caches/{bid}"),
                      ignore_errors=True)
    try:
        with open(marca, "w") as f:
            f.write(str(VERSION))
    except Exception:
        pass


_LOCK = None    # referencia viva al candado de instancia única (no cerrar)


def _instancia_unica():
    """True si esta es la única instancia de Gestiq. Dos procesos a la vez
    pelean por el perfil del WebView (WebView2 lo bloquea en Windows) y la
    segunda ventana queda congelada; mejor no abrirla. Al morir el proceso el
    sistema libera el candado solo. Si el candado no se puede crear, se deja
    pasar (nunca bloquear el arranque por esto)."""
    global _LOCK
    import tempfile
    ruta = os.path.join(tempfile.gettempdir(), "gestiq_instancia.lock")
    try:
        f = open(ruta, "a+")
    except Exception:
        return True
    try:
        if sys.platform.startswith("win"):
            import msvcrt
            msvcrt.locking(f.fileno(), msvcrt.LK_NBLCK, 1)
        else:
            import fcntl
            fcntl.flock(f, fcntl.LOCK_EX | fcntl.LOCK_NB)
    except Exception:
        try: f.close()
        except Exception: pass
        if sys.platform.startswith("win"):
            try:                            # feedback: sin esto parecía "no abre"
                import ctypes
                ctypes.windll.user32.MessageBoxW(
                    None, "Gestiq ya está abierto (revisa la barra de tareas).",
                    "Gestiq", 0x50040)      # info + siempre visible + al frente
            except Exception:
                pass
        return False                        # otra instancia tiene el candado
    _LOCK = f
    return True


def _rematar(api, salir=os._exit):
    """Último paso del cierre: espera (acotado) a que el autoguardado de las
    corridas termine de escribir y MATA el proceso. Se llama desde dos sitios
    porque webview.start() a veces NO retorna tras cerrar la ventana (visto en
    Windows): el proceso quedaba vivo, retenía el candado de instancia única y
    el siguiente arranque "no abría". Con esto el proceso muere sí o sí."""
    api._closing = True
    for t in api._tabs.values():
        th = t._thread
        if th is not None and th.is_alive():
            th.join(timeout=15)
    salir(0)


def main():
    if not _instancia_unica():
        return                              # ya hay un Gestiq abierto
    if sys.platform.startswith("win"):
        try:                                # DPI por monitor: texto nítido y sin
            import ctypes                   # reescalado (si ya estaba fijada por
            ctypes.windll.shcore.SetProcessDpiAwareness(2)   # pywebview, falla y ya)
        except Exception:
            pass
        # OJO: NO pasar flags por WEBVIEW2_ADDITIONAL_BROWSER_ARGUMENTS.
        # Se probó (--disable-features=..., --ignore-gpu-blocklist, etc.) y
        # SALIÓ MAL: --disable-features REEMPLAZA la lista interna del motor
        # (no se suma) y dejó el compositor congelado a medio arranque en el
        # equipo de Toxic. El render sano no necesita ayudas: la causa real de
        # las "animaciones muertas" era un @media prefers-reduced-motion del
        # CSS (ya eliminado), no la GPU.
        os.environ.pop("WEBVIEW2_ADDITIONAL_BROWSER_ARGUMENTS", None)
    _limpiar_cache_web()
    api = Api()
    kwargs = dict(
        title="Gestiq", url=_ruta("gestiq_ui.html"), js_api=api,
        width=1140, height=780, min_size=(960, 660),
        background_color="#101018",
    )
    try:
        api._win = webview.create_window(vibrancy=True, **kwargs)   # blur nativo (solo macOS)
    except TypeError:
        api._win = webview.create_window(**kwargs)

    def al_cerrar():
        corriendo = [t for t in api._tabs.values() if t._running]
        if corriendo:
            seguir = True
            try:
                seguir = bool(api._win.create_confirmation_dialog(
                    "Consulta en curso",
                    "Hay una consulta en curso; si cierras ahora se detendrá.\n"
                    "Lo ya consultado queda en el archivo de autoguardado.\n\n"
                    "¿Cerrar de todas formas?"))
            except Exception:
                pass                       # sin soporte de diálogo → cerrar como antes
            if not seguir:
                return False               # cancela el cierre
        api._closing = True                # desde aquí ya no se pinta nada en la UI
        for t in corriendo:
            try: t._do_stop()              # cancela la tarea → _runner autoguarda
            except Exception: pass
        try: updater.aplicar_y_reiniciar()   # si hay update listo, reemplaza y relanza
        except Exception: pass
        # Remate en segundo plano: aunque webview.start() no retorne (pasa a
        # veces en Windows), el proceso muere igual tras el autoguardado.
        threading.Thread(target=lambda: (time.sleep(2), _rematar(api)),
                         daemon=True).start()
        return True
    api._win.events.closing += al_cerrar

    # Auto-actualización silenciosa en segundo plano.
    def _aviso_update(tag):
        api.js_toast("ok", "Actualización",
                     f"Gestiq {tag} se instalará al cerrar la aplicación.")
    try:
        updater.iniciar_en_segundo_plano(on_listo=_aviso_update)
    except Exception:
        pass

    # private_mode=False conserva el tema elegido. En Windows se fija además un
    # perfil web PROPIO y estable: sin esto WebView2 usa la carpeta compartida
    # de pywebview y un perfil bloqueado/corrupto congela la ventana al abrir.
    inicio = dict(private_mode=False)
    if sys.platform.startswith("win"):
        base = os.environ.get("LOCALAPPDATA") or os.path.expanduser("~")
        perfil = os.path.join(base, "Gestiq", "webview")
        try: os.makedirs(perfil, exist_ok=True)
        except Exception: perfil = None
        if perfil: inicio["storage_path"] = perfil
    try:
        webview.start(**inicio)
    except TypeError:
        try: webview.start(private_mode=False)
        except TypeError: webview.start()

    # Ventana cerrada y webview.start() retornó: rematar ya mismo (el watchdog
    # de al_cerrar cubre el caso en que start() no retorna). El ayudante del
    # updater ya quedó lanzado como proceso aparte y sobrevive a esta salida.
    _rematar(api)


if __name__ == "__main__":
    main()
