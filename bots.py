# -*- coding: utf-8 -*-
"""
bots.py — Lógica viva de Gestiq: bots IMAGINE + GUARDIAN, Excel y preferencias.

SIN ninguna dependencia de interfaz (ni tkinter ni CustomTkinter): la capa de
UI (gestiq_web.py) hereda de estos bots y sobreescribe los hooks de interfaz
(after, _upd, _ask_ready, _on_finish, _set_running) y el shim `messagebox`.

Extraído de gestiq.py (refactor 02-jul-2026) para:
  - que la app web no dependa de Tk (gestiq.py hacía sys.exit sin CustomTkinter),
  - eliminar el problema de caché .pyc sirviendo lógica vieja,
  - aligerar el ejecutable compilado.
gestiq.py queda solo como referencia de la UI vieja; NO editar lógica allí.
"""

import os, re, json, time, base64, threading, asyncio
from urllib.parse import unquote

from copy import copy
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter, column_index_from_string

try:
    from playwright.async_api import async_playwright, TimeoutError as PWTimeout
    HAVE_PW = True
except ImportError:
    HAVE_PW = False

try:
    import licencia
except Exception:
    licencia = None

# ── Preferencia "Recordar inicio de sesión" (IMAGINE/GUARDIÁN) ───────────────
def nav_recordar():
    """¿Recordar la sesión del navegador entre corridas? (default sí)."""
    try:
        return licencia is None or licencia.recordar_get()
    except Exception:
        return True

def nav_borrar_todo():
    """Olvida las sesiones guardadas de IMAGINE y GUARDIÁN (al cambiar el switch)."""
    for mod in ("imagine", "guardian"):
        try:
            os.remove(os.path.join(os.path.expanduser("~"), f".gestiq_nav_{mod}"))
        except OSError:
            pass

# Web de registro y pago
REGISTRO_URL = "https://toxicamazonas-svg.github.io/gestiq/cuenta.html"

IMAGINE_BASE = "https://imaginex-intranet.segurosbolivar.com"
GUARDIAN_API = "https://gestion-api.guardiandelaproductividad.com"


# ── Shim de mensajes (la UI lo reemplaza por toasts) ─────────────────────────
class _MsgNull:
    def showinfo(self, *a, **k):    pass
    def showwarning(self, *a, **k): pass
    def showerror(self, *a, **k):   pass
    def askyesno(self, *a, **k):    return True

messagebox = _MsgNull()


# ── Preferencias locales por cuenta (nombre, foto, tema, módulo inicial) ─────
PREFS_PATH = os.path.join(os.path.expanduser("~"), ".gestiq_prefs.json")

def _prefs_all():
    try:
        with open(PREFS_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def _prefs_get(email):
    return _prefs_all().get((email or "").lower(), {})

def _prefs_set(email, **kw):
    """Guarda preferencias de una cuenta. Un valor None elimina la clave."""
    email = (email or "").lower()
    if not email:
        return
    data = _prefs_all()
    cur = data.get(email, {})
    cur.update(kw)
    data[email] = {k: v for k, v in cur.items() if v is not None}
    try:
        with open(PREFS_PATH, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False)
    except Exception:
        pass


# ── Estilos Excel (mismos colores/formato del archivo original) ─────────────
def xfill(rgb): return PatternFill(start_color=rgb, end_color=rgb, fill_type='solid')
F_GREEN   = xfill('00B050')   # verde sólido (APROBADO)
F_GREENLT = xfill('92D050')   # verde claro (observaciones autorizadas)
F_RED     = xfill('FF0000')   # rojo (devuelto / rechazado)
F_ORANGE  = xfill('FFC000')   # naranja (no requiere / errado)
F_YELLOW  = xfill('FFFF00')   # amarillo típico de Excel (no requiere aprobación)
F_BLACK   = xfill('000000')   # negro con letra blanca (no encontrada / pendiente)

AL_CELL = Alignment(horizontal='center', vertical='center', wrap_text=True)

def style_cell(cell, fill=None, white=False):
    """Aplica el formato del Excel original: Calibri 7 negrita, centrado, ajustar texto."""
    if fill: cell.fill = fill
    cell.font      = Font(name='Calibri', size=7, bold=True,
                          color='FFFFFF' if white else '000000')
    cell.alignment = AL_CELL


# ── Estilos personalizables de los resultados (Preferencias → Resultados) ───
# Los defaults REPLICAN el estilo clásico de siempre; el usuario puede cambiar
# fondo/letra por tipo de resultado y fuente/tamaño/negrita globales. La capa
# web estampa la personalización en tab.estilos al iniciar cada corrida.
ESTILOS_DEF = {
    "fuente":  {"nombre": "Calibri", "tam": 7, "negrita": True},
    "imagine": {"ok":    {"fondo": "00B050", "letra": "000000"},
                "dev":   {"fondo": "FF0000", "letra": "000000"},
                "err":   {"fondo": "FFC000", "letra": "000000"},
                "nota":  {"fondo": "FF0000", "letra": "000000"},
                "otros": {"fondo": "",       "letra": "000000"}},
    "guardian":{"aprobado":  {"fondo": "00B050", "letra": "000000"},
                "rechazado": {"fondo": "FF0000", "letra": "FFFFFF"},
                "pendiente": {"fondo": "FF0000", "letra": "000000"},
                "noenc":     {"fondo": "000000", "letra": "FFFFFF"},
                "noinf":     {"fondo": "FFC000", "letra": "000000"},
                "noreq":     {"fondo": "FFFF00", "letra": "000000"},
                "otros":     {"fondo": "",       "letra": "000000"}},
}


def _hex6(v, defecto):
    """Valida un color hex de 6 dígitos; si no sirve, cae al default."""
    v = str(v or "").strip().lstrip("#").upper()
    return v if re.fullmatch(r"[0-9A-F]{6}", v) else defecto


# ════════════════════════════════════════════════════════════════════════════
#  Base común de los bots
# ════════════════════════════════════════════════════════════════════════════
class BaseBot:
    """Orquestación compartida: hilo/loop asyncio, stop, autoguardado y
    helpers de Excel. Los hooks de UI son no-op aquí; la capa web los tapa."""

    AUTOSAVE_CADA = 10          # autoguarda cada N filas procesadas
    autosave_on = True          # preferencia del usuario (la capa web la fija por corrida)
    estilos = None              # personalización de resultados (la fija la capa web)

    def pintar(self, cell, modulo, clave):
        """Formatea una celda de resultado con la personalización del usuario
        (fondo/letra por tipo + fuente/tamaño/negrita globales); sin
        personalización replica el estilo clásico (ESTILOS_DEF)."""
        conf = self.estilos or {}
        base = ESTILOS_DEF.get(modulo, {}).get(clave, {})
        e = dict(base); e.update((conf.get(modulo) or {}).get(clave) or {})
        f = dict(ESTILOS_DEF["fuente"]); f.update(conf.get("fuente") or {})
        fondo = _hex6(e.get("fondo"), base.get("fondo", "")) if e.get("fondo") else ""
        letra = _hex6(e.get("letra"), base.get("letra", "000000") or "000000")
        if fondo:
            cell.fill = xfill(fondo)
        try:    tam = max(5.0, min(20.0, float(f.get("tam") or 7)))
        except Exception: tam = 7.0
        cell.font = Font(name=str(f.get("nombre") or "Calibri")[:40], size=tam,
                         bold=bool(f.get("negrita", True)), color=letra)
        cell.alignment = AL_CELL

    # Estado (la capa web lo inicializa en init_comun; defaults por seguridad)
    key = ""
    xl_path = None
    wb = None
    _stop = False
    _running = False
    _loop = _task = _thread = None
    _login_ev = None

    # ── Hooks de UI (sobreescritos por la capa web) ──
    def after(self, _ms, fn=None, *a):
        if fn is None: return
        try: fn(*a)
        except Exception: pass

    def _upd(self, cur, tot, msg=""): pass
    def _set_prog(self, v): pass
    def _set_running(self, r): self._running = r
    def _on_finish(self): self._set_running(False)

    def _ask_ready(self, event, system_name):
        event.set()             # sin UI no hay login manual que esperar

    # ── Ejecución / parada ──
    def _runner(self):
        """Ejecuta la automatización en su propio event loop, cancelable al instante."""
        loop = asyncio.new_event_loop()
        self._loop = loop
        asyncio.set_event_loop(loop)
        self._autosave_dirty = False
        self._autosave_avisado = False
        self.resumen = None                         # resumen final para la UI
        self._res = {}
        self._task = loop.create_task(self._automate())
        try:
            loop.run_until_complete(self._task)
        except asyncio.CancelledError:
            self._autosave()                        # detenido: lo consultado queda en disco
        except Exception as e:
            self.app.log(self, f"Error inesperado: {e}", "error")
        finally:
            try:
                loop.run_until_complete(loop.shutdown_asyncgens())
            except Exception:
                pass
            loop.close()
            self._loop = self._task = None

    def _do_stop(self):
        """Detiene la búsqueda de inmediato: aborta el caso en curso y cierra el navegador."""
        self._stop = True
        ev = self._login_ev
        if ev is not None and not ev.is_set():
            ev.set()        # _automate ve _stop=True y cierra el navegador ordenadamente
        else:
            # Cancela la tarea asyncio: aborta la petición en curso sin esperar
            loop, task = self._loop, self._task
            if loop is not None and task is not None and not task.done():
                loop.call_soon_threadsafe(task.cancel)
        self.app.log(self, "Detenido por el usuario. Lo ya consultado quedó en el Excel.", "warn")

    # ── Autoguardado (protege lo consultado ante cierres, fallos o crashes) ──
    def _autosave_path(self):
        base, _ = os.path.splitext(self.xl_path or "resultado")
        mod = str(getattr(self, "key", "") or self.__class__.__name__).upper()
        return f"{base} (autoguardado {mod}).xlsx"

    def _autosave(self, force=True):
        """Copia de respaldo junto al archivo original. Solo escribe si hay
        resultados nuevos. Los autoguardados periódicos (force=False) se
        espacian mínimo 30 s: guardar un libro grande tarda segundos y
        hacerlo muy seguido hace sentir la app trabada."""
        if getattr(self, "autosave_on", True) is False:
            return              # desactivado por el usuario en Preferencias
        if not self.wb or not getattr(self, "_autosave_dirty", False):
            return
        if not force and time.time() - getattr(self, "_autosave_t", 0) < 30:
            return
        try:
            p = self._autosave_path()
            self.wb.save(p)
            self._autosave_dirty = False
            self._autosave_t = time.time()
            if not getattr(self, "_autosave_avisado", False):
                self._autosave_avisado = True
                self.app.log(self, f"Autoguardado activo: {os.path.basename(p)}", "info")
        except Exception as e:
            self.app.log(self, f"No se pudo autoguardar: {e}", "warn")

    # ── Resumen final (conteo por categoría, para la tarjeta de la UI) ──────
    def _res_add(self, etiqueta, clase):
        if not hasattr(self, "_res"):
            self._res = {}
        n, _ = self._res.get(etiqueta, (0, clase))
        self._res[etiqueta] = (n + 1, clase)

    def _res_fin(self, hecho):
        self.resumen = {
            "total": hecho,
            "errores": self._res.get("Errores", (0, ""))[0],
            "items": [{"t": t, "n": n, "c": c} for t, (n, c) in self._res.items()],
        }

    # ── Sesión del navegador recordada (storage_state de Playwright) ────────
    # Evita loguearse en IMAGINE/GUARDIÁN en cada corrida. El archivo se
    # guarda ofuscado con la llave del equipo (misma técnica que licencia).
    # La preferencia "Recordar inicio de sesión" (switch en Preferencias)
    # gobierna TODO el mecanismo: apagada, ni se carga ni se guarda nada.
    def _nav_path(self):
        mod = (self.key or self.__class__.__name__).lower()
        return os.path.join(os.path.expanduser("~"), f".gestiq_nav_{mod}")

    def _nav_guardar(self, estado):
        if not nav_recordar():
            return
        try:
            data = json.dumps(estado).encode()
            if licencia is not None:
                data = b"GQ1" + licencia._ofusca(data)
            else:
                data = b"B64" + base64.b64encode(data)
            with open(self._nav_path(), "wb") as f:
                f.write(data)
            try: os.chmod(self._nav_path(), 0o600)
            except Exception: pass
        except Exception:
            pass

    def _nav_cargar(self):
        if not nav_recordar():
            return None
        try:
            raw = open(self._nav_path(), "rb").read()
            if raw.startswith(b"GQ1") and licencia is not None:
                return json.loads(licencia._ofusca(raw[3:]))
            if raw.startswith(b"B64"):
                return json.loads(base64.b64decode(raw[3:]))
        except Exception:
            pass
        return None

    # ── Ventana del navegador del bot ────────────────────────────────────────
    # Con la sesión recordada el trabajo corre por API y la ventana blanca de
    # Chromium solo estorba: se abre minimizada. Si hace falta login manual,
    # _nav_mostrar la trae de vuelta. CDP solo existe en Chromium (nuestro caso).
    async def _nav_minimizar(self, ctx, page):
        try:
            cdp = await ctx.new_cdp_session(page)
            w = await cdp.send("Browser.getWindowForTarget")
            await cdp.send("Browser.setWindowBounds",
                           {"windowId": w["windowId"],
                            "bounds": {"windowState": "minimized"}})
            await cdp.detach()
            self.app.log(self, "Navegador minimizado — el trabajo sigue en segundo plano.", "info")
        except Exception:
            pass                # sin CDP la ventana queda visible: inofensivo

    async def _nav_mostrar(self, ctx, page):
        try:
            cdp = await ctx.new_cdp_session(page)
            w = await cdp.send("Browser.getWindowForTarget")
            await cdp.send("Browser.setWindowBounds",
                           {"windowId": w["windowId"],
                            "bounds": {"windowState": "normal"}})
            await cdp.detach()
            await page.bring_to_front()
        except Exception:
            pass

    # ── Helpers Excel ────────────────────────────────────────────────────────
    def _find_header_row(self, ws):
        for i, row in enumerate(ws.iter_rows(max_row=10), 1):
            if sum(1 for c in row if c.value) >= 3:
                return i
        return 1

    def _find_col(self, ws, names, hdr_row=None, last=False):
        hr    = hdr_row or self._find_header_row(ws)
        found = None
        for cell in ws[hr]:
            if cell.value and str(cell.value).strip().upper() in [n.upper() for n in names]:
                if not last: return cell.column
                found = cell.column
        return found

    def _get_or_create_col(self, ws, name, hdr_row, last=False):
        col = self._find_col(ws, [name], hdr_row, last=last)
        if col: return col
        # OJO: no usar ws.max_column (el formato fantasma de la plantilla lo
        # infla y la columna nueva quedaría lejos de la tabla): usar la última
        # celda del encabezado con texto.
        ult = 0
        for cell in ws[hdr_row]:
            if cell.value is not None and str(cell.value).strip() != "":
                ult = max(ult, cell.column)
        col = (ult or ws.max_column) + 1
        ws.cell(row=hdr_row, column=col, value=name)
        return col

    async def _automate(self): pass


# ════════════════════════════════════════════════════════════════════════════
#  Bot IMAGINE
# ════════════════════════════════════════════════════════════════════════════
class ImagineBot(BaseBot):

    # Palabras que definen el ESTADO real del caso. Lista ÚNICA para categoría,
    # color de celda y para distinguir la observación de estado de las notas
    # adicionales (p.ej. "Recuerde que las firmas… se devolverá" NO es estado:
    # "devolverá" no contiene "devuelve" — no agregar variantes que choquen).
    KW_OK  = ("autorizado", "aprobado", "pre factura")
    KW_DEV = ("devuelve", "devuelto", "rechaza", "no coincide")

    # ── Categoría del resultado (para el resumen final) ─────────────────────
    @staticmethod
    def _categoria(result):
        if result is None:
            return ("Pendientes", "warn")
        r = result.lower()
        if result.startswith("ERROR"):
            return ("Errores", "err")
        if "no encontrado" in r:
            return ("No encontrados", "warn")
        if "errado" in r:
            return ("Errados", "warn")
        if any(x in r for x in ImagineBot.KW_OK):
            return ("Aprobados", "ok")
        if any(x in r for x in ImagineBot.KW_DEV):
            return ("Rechazados", "err")
        return ("Otros", "info")

    # ── Columna NOTAS pegada a la salida (SIN pisar lo que haya al lado) ─────
    def _col_notas(self, ws, hdr, out_c):
        """Devuelve la columna para las notas adicionales (out_c+1). Si esa
        posición ya tiene OTRO encabezado (p.ej. FECHA FACTURACION), inserta
        una columna nueva ahí y corre lo demás una casilla a la derecha —
        openpyxl mueve celdas y estilos pero NO anchos/combinadas/filtro:
        se corren a mano (verificado 03-jul-2026)."""
        col = out_c + 1
        cab = ws.cell(row=hdr, column=col)
        txt = str(cab.value).strip().upper() if cab.value else ""
        if txt == "NOTAS":
            return col                              # ya existe (corrida anterior)
        if txt:                                     # hay otra columna: INSERTAR
            ws.insert_cols(col)
            dims = ws.column_dimensions
            movidas = []                            # anchos/ocultas → una a la derecha
            for letra in list(dims):
                i = column_index_from_string(letra)
                if i >= col:
                    d = dims.pop(letra)
                    movidas.append((i + 1, d.width, d.hidden))
            for i, w, hid in movidas:
                nd = dims[get_column_letter(i)]
                if w is not None: nd.width = w
                nd.hidden = hid
            for rng in list(ws.merged_cells.ranges):    # combinadas
                if rng.min_col >= col:   rng.shift(col_shift=1)
                elif rng.max_col >= col: rng.max_col += 1
            try:                                        # rango del autofiltro
                ref = ws.auto_filter.ref
                if ref:
                    from openpyxl.worksheet.cell_range import CellRange
                    r = CellRange(ref)
                    if r.min_col >= col:   r.shift(col_shift=1)
                    elif r.max_col >= col: r.max_col += 1
                    ws.auto_filter.ref = r.coord
            except Exception:
                pass
            dims[get_column_letter(col)].width = 45
            cab = ws.cell(row=hdr, column=col)
        cab.value = "NOTAS"
        try: cab._style = copy(ws.cell(row=hdr, column=out_c)._style)
        except Exception: pass
        return col

    # ── Orquestador (multi-hoja en una sola corrida) ─────────────────────────
    async def _automate(self):
        hojas = [h for h in (getattr(self, "hojas", None) or [self.v_sheet.get()]) if h]

        # Plan por hoja: columnas y filas pendientes
        planes = []
        for nombre in hojas:
            if nombre not in self.wb.sheetnames:
                self.app.log(self, f"Hoja «{nombre}» no existe — omitida.", "warn"); continue
            ws  = self.wb[nombre]
            hdr = self._find_header_row(ws)
            cc  = self._find_col(ws, ["CASO", self.v_caso.get().strip()], hdr)
            if not cc:
                self.app.log(self, f"Hoja «{nombre}»: columna CASO no encontrada — omitida.", "error")
                continue
            out_c = self._get_or_create_col(ws, self.v_out.get().strip().upper(), hdr)
            jobs = []
            for row in ws.iter_rows(min_row=hdr + 1):
                cv = str(row[cc - 1].value).strip() if row[cc - 1].value else ""
                ov = str(ws.cell(row=row[0].row, column=out_c).value).strip().upper() \
                     if ws.cell(row=row[0].row, column=out_c).value else ""
                if cv and cv.upper() not in ("NONE", "NAN"):
                    # Vacía, PENDIENTE o ERROR (los errores de red se reintentan
                    # en la siguiente corrida sin borrar la celda a mano).
                    if not ov or ov == "PENDIENTE" or ov.startswith("ERROR"):
                        jobs.append((row[0].row, cv))
            if jobs:
                planes.append((nombre, ws, hdr, out_c, jobs))
            else:
                self.app.log(self, f"Hoja «{nombre}»: sin casos pendientes.", "info")

        total = sum(len(p[4]) for p in planes)
        if not total:
            self.app.log(self, "No hay casos pendientes para procesar.", "warn")
            self.after(0, self._on_finish); return

        multi = len(planes) > 1
        self.app.log(self, f"{total} casos a procesar" +
                     (f" en {len(planes)} hojas." if multi else "."), "info")
        not_found = []
        hecho = 0

        async with async_playwright() as pw:
            browser = await pw.chromium.launch(headless=False, slow_mo=250)
            estado  = self._nav_cargar()
            try:
                ctx = await browser.new_context(storage_state=estado) if estado \
                      else await browser.new_context()
            except Exception:
                ctx = await browser.new_context()
            page = await ctx.new_page()
            if estado:                  # sesión recordada: la ventana no hace falta
                await self._nav_minimizar(ctx, page)
            await page.goto(IMAGINE_BASE)

            # ¿Sigue viva la sesión recordada? (sondeo barato; si no, login manual)
            viva = False
            if estado:
                try:
                    await self._ajax(ctx, opcion="cargaInfoRadicacion",
                                     na="1", cons="si", band="0")
                    viva = True
                    self.app.log(self, "Sesión de IMAGINE recordada — sin login manual.", "ok")
                except Exception:
                    self.app.log(self, "La sesión guardada expiró — inicia sesión de nuevo.", "info")

            if not viva:
                await self._nav_mostrar(ctx, page)      # el login manual sí necesita la ventana
                ev = threading.Event()
                self.after(0, lambda: self._ask_ready(ev, "Imagine (solo inicia sesión — el bot navega solo)"))
                ev.wait()
                if self._stop:
                    await browser.close(); self.after(0, self._on_finish); return
                try: self._nav_guardar(await ctx.storage_state())
                except Exception: pass
                await self._nav_minimizar(ctx, page)    # login listo: fuera del medio

            corte = False
            nota_cols = {}                      # hoja → columna NOTAS (creada al 1er uso)
            for nombre, ws, hdr, out_c, jobs in planes:
                if self._stop or corte: break
                if multi:
                    self.app.log(self, f"— Hoja «{nombre}» ({len(jobs)} casos) —", "info")
                for rn, caso in jobs:
                    if self._stop: break
                    hecho += 1
                    pre = f"[{nombre}] " if multi else ""
                    self.after(0, lambda h=hecho, c=caso, p=pre: (
                        self._upd(h, total, f"{p}Caso {c}"),
                        self.app.log(self, f"→ Caso {c}", "info")
                    ))
                    try:
                        result, extra = await self._procesar_caso(ctx, caso)
                    except Exception as e:
                        if "Sesión expirada" in str(e):
                            self.app.log(self, "Sesión expirada — proceso detenido. "
                                               "Vuelve a iniciar sesión y reinicia.", "error")
                            hecho -= 1
                            corte = True
                            break
                        result, extra = f"ERROR: {str(e)[:80]}", ""
                        self.app.log(self, str(e), "error")

                    self._res_add(*self._categoria(result))
                    if result is None:
                        self.app.log(self, f"  {caso}: PENDIENTE — omitido", "warn")
                        continue

                    cell = ws.cell(row=rn, column=out_c)
                    cell.value = result
                    self._autosave_dirty = True
                    rl = result.lower()
                    if any(x in rl for x in self.KW_OK):
                        self.pintar(cell, "imagine", "ok")
                    elif any(x in rl for x in self.KW_DEV):
                        self.pintar(cell, "imagine", "dev")
                    elif any(x in rl for x in ["errado", "no encontrado", "error"]):
                        self.pintar(cell, "imagine", "err")
                        if "no encontrado" in rl: not_found.append(caso)
                    else:
                        self.pintar(cell, "imagine", "otros")

                    # Nota(s) adicionales de la misma gestión → columna NOTAS
                    # pegada a la salida (se inserta sin pisar lo que haya).
                    if extra:
                        if nombre not in nota_cols:
                            nota_cols[nombre] = self._col_notas(ws, hdr, out_c)
                        c2 = ws.cell(row=rn, column=nota_cols[nombre], value=extra)
                        self.pintar(c2, "imagine", "nota")

                    self.app.log(self, f"  {caso}: {result[:80]}",
                                 "error" if "no encontrado" in rl else "ok")
                    if extra:
                        self.app.log(self, f"  {caso}: nota adicional → {extra[:70]}", "warn")
                    if hecho % self.AUTOSAVE_CADA == 0:
                        self._autosave(force=False)

            self._autosave()
            try: self._nav_guardar(await ctx.storage_state())
            except Exception: pass
            await browser.close()

        if not_found:
            self.app.log(self, f"No encontrados: {', '.join(not_found)}", "warn")
        self._res_fin(hecho)
        self.app.log(self, f"Listo. {hecho}/{total} procesados.", "ok")
        self.after(0, self._on_finish)

    # ── Lógica de un caso (protocolo AJAX directo — verificado en el sitio) ──
    # cargaFuncion.php es solo un "cascarón" que carga todo por AJAX; el bot
    # llama directamente los 3 endpoints (POST /arp/radicador/ajax/_radicador.php):
    #   1. opcion=cargaInfoRadicacion&na=<caso>&cons=si&band=0
    #      → info del caso (No. Caso, Estado, Razón Empresa)
    #      → si no existe: "No hay resultados para mostrar..."
    #   2. opcion=cargaHistGestion&na=<caso>&adic=0&band=0
    #      → tabla de actividades table-striped [Actividad|Usuario|Estado|...]
    #   3. opcion=cargaObsGestionTotal&na=<caso>
    #      → tabla de observaciones editorTip [Fecha|Usuario|Actividad|Observacion|Adjunto]
    # La sesión viaja en las cookies del contexto de Playwright (login manual).
    @staticmethod
    def _strip(html):
        return " ".join(re.sub(r"<[^>]+>", " ", html).split())

    @staticmethod
    def _filas(html_tabla):
        """Devuelve las filas <tr> de una tabla como listas de celdas (texto plano)."""
        filas = []
        for tr in re.findall(r"<tr[^>]*>(.*?)</tr>", html_tabla, re.S | re.I):
            tds = re.findall(r"<td[^>]*>(.*?)</td>", tr, re.S | re.I)
            filas.append([ImagineBot._strip(td) for td in tds])
        return filas

    async def _ajax(self, ctx, **datos):
        r = await ctx.request.post(
            f"{IMAGINE_BASE}/arp/radicador/ajax/_radicador.php",
            form=datos, timeout=30000)
        html = await r.text()
        low = self._strip(html).lower()
        # Sesión expirada → el servidor devuelve la página de login
        if "ingrese su usuario" in low or "autenticaci" in low:
            raise RuntimeError("Sesión expirada — vuelve a iniciar sesión en el navegador")
        return html

    async def _procesar_caso(self, ctx, caso):
        """Devuelve (resultado, extra): resultado va en la columna de salida y
        extra (notas adicionales de la misma gestión) en la celda de al lado."""
        caso = str(caso).split(".")[0].strip()

        # 1) Info del caso
        html = await self._ajax(ctx, opcion="cargaInfoRadicacion",
                                na=caso, cons="si", band="0")
        body = self._strip(html)

        if "No hay resultados" in body or f"No. Caso: {caso}" not in body:
            return "CASO NO ENCONTRADO", ""

        # Empresa: texto entre "Razón Empresa:" y "Nit/CC"
        # (\S*n tolera problemas de codificación en la tilde de "Razón")
        m = re.search(r"Raz\S*n Empresa:\s*(.*?)\s*Nit/CC", body)
        empresa = m.group(1).upper() if m else body.upper()
        if "IPRECON" not in empresa:
            return "CASO ERRADO", ""

        # 2) Actividades: tabla class="table table-striped..." →
        # [Actividad | Usuario | Estado | Fecha Asig. | Fecha Cierre | Gestionar]
        html_act = await self._ajax(ctx, opcion="cargaHistGestion",
                                    na=caso, adic="0", band="0")
        mt = re.search(r'<table[^>]*table-striped[^>]*>(.*?)</table>', html_act, re.S | re.I)
        if mt:
            for celdas in self._filas(mt.group(1)):
                if len(celdas) >= 3 and "pendiente" in celdas[2].lower():
                    return None, ""                  # actividad pendiente → omitir

        # 3) Observaciones (mismo endpoint que usa el botón del modal)
        html2 = await self._ajax(ctx, opcion="cargaObsGestionTotal", na=caso)

        # Tabla editorTip: [Fecha | Usuario | Actividad | Observacion | Adjunto]
        mo = re.search(r'<table[^>]*editorTip[^>]*>(.*?)</table>', html2, re.S | re.I)
        obs = []                                     # [(fecha, texto)] en orden de tabla
        if mo:
            for celdas in self._filas(mo.group(1)):
                if len(celdas) < 4:
                    continue
                fecha, texto = celdas[0], celdas[3]
                # solo filas cuya 1ª celda es fecha real (salta el encabezado)
                if texto and re.match(r"\d{4}-\d{2}-\d{2}", fecha):
                    obs.append((fecha, texto))

        if not obs:
            return "SIN OBSERVACION", ""

        # ESTADO = la observación MÁS RECIENTE que diga aprobado/prefactura o
        # devuelto/rechazado (KW_OK/KW_DEV); así una nota extra (p.ej. "Recuerde
        # que las firmas…") no tapa el estado real. Las demás notas del MISMO
        # DÍA en adelante (misma gestión) se devuelven unidas como extra.
        estado = None
        for fecha, texto in obs:
            tl = texto.lower()
            if any(k in tl for k in self.KW_OK + self.KW_DEV) \
               and (estado is None or fecha >= estado[0]):
                estado = (fecha, texto)

        if estado is None:                           # sin estado claro: como siempre
            return max(obs, key=lambda o: o[0])[1], ""

        dia = estado[0][:10]
        vistos, extras = {estado[1]}, []
        for fecha, texto in obs:
            if fecha[:10] >= dia and texto not in vistos:
                extras.append(texto); vistos.add(texto)
        return estado[1], " | ".join(extras)


# ════════════════════════════════════════════════════════════════════════════
#  Bot GUARDIAN
# ════════════════════════════════════════════════════════════════════════════
class GuardianBot(BaseBot):

    # ── Categoría del resultado (para el resumen final) ─────────────────────
    @staticmethod
    def _categoria(result):
        ru = (result or "").upper()
        if ru.startswith("ERROR"):
            return ("Errores", "err")
        if "NO ENCONTR" in ru or "ACTIVIDAD NO" in ru:
            return ("No encontrados", "warn")
        if "RECHAZADO" in ru:
            return ("Rechazados", "err")
        if "PENDIENTE" in ru:
            return ("Pendientes", "warn")
        if "APROBADO" in ru:
            return ("Aprobados", "ok")
        if "NO HAY INFORME" in ru:
            return ("No hay informe (revisar)", "warn")
        if "NO NECESITA" in ru:               # texto viejo (archivos anteriores)
            return ("No necesita aprobación", "info")
        return ("Otros", "info")

    # ── Guardián: extraer token/usuario de las cookies del contexto ─────────
    @staticmethod
    def _tok_uid(cookies):
        tok = unquote(cookies.get("sesion", ""))
        uid = unquote(cookies.get("user_id", ""))
        # la cookie user_id viene en base64 (p.ej. "NDMyMDM=" → "43203")
        if uid and not uid.isdigit():
            try:    uid = base64.b64decode(uid).decode().strip()
            except Exception: pass
        return tok, uid

    # ── Orquestador (multi-hoja en una sola corrida) ─────────────────────────
    async def _automate(self):
        hojas = [h for h in (getattr(self, "hojas", None) or [self.v_sheet.get()]) if h]

        # Plan por hoja: columnas y filas pendientes
        planes = []
        for nombre in hojas:
            if nombre not in self.wb.sheetnames:
                self.app.log(self, f"Hoja «{nombre}» no existe — omitida.", "warn"); continue
            ws  = self.wb[nombre]
            hdr = self._find_header_row(ws)
            cc  = self._find_col(ws, ["CRONOGRAMA", self.v_cron.get().strip()], hdr)
            sc  = self._find_col(ws, ["SECUENCIA",  self.v_sec.get().strip()],  hdr)
            if not cc or not sc:
                self.app.log(self, f"Hoja «{nombre}»: columnas CRONOGRAMA o SECUENCIA "
                                   "no encontradas — omitida.", "error")
                continue
            out_c = self._get_or_create_col(ws, self.v_out.get().strip().upper(), hdr)
            self.app.log(self, f"Hoja «{nombre}»: encabezado fila {hdr} | CRON col {cc} | "
                               f"SEC col {sc} | salida col {out_c}", "info")
            jobs = []
            for row in ws.iter_rows(min_row=hdr + 1):
                cv = str(row[cc - 1].value).strip() if row[cc - 1].value else ""
                sv = str(row[sc - 1].value).strip() if row[sc - 1].value else ""
                ov = str(ws.cell(row=row[0].row, column=out_c).value).strip().upper() \
                     if ws.cell(row=row[0].row, column=out_c).value else ""
                if cv and sv and cv.upper() not in ("NONE", "NAN", ""):
                    # Vacía, PENDIENTE o ERROR (se reintentan en la siguiente corrida)
                    if not ov or "PENDIENTE" in ov or ov.startswith("ERROR"):
                        jobs.append((row[0].row, cv, sv))
            if jobs:
                planes.append((nombre, ws, out_c, jobs))
            else:
                self.app.log(self, f"Hoja «{nombre}»: sin filas pendientes.", "info")

        total = sum(len(p[3]) for p in planes)
        if not total:
            self.app.log(self, "No hay filas pendientes para procesar.", "warn")
            self.after(0, self._on_finish); return

        multi = len(planes) > 1
        self.app.log(self, f"{total} filas a procesar" +
                     (f" en {len(planes)} hojas." if multi else "."), "info")
        no_resultado = []
        hecho = 0

        async with async_playwright() as pw:
            browser = await pw.chromium.launch(headless=False, slow_mo=200)
            estado  = self._nav_cargar()
            try:
                ctx = await browser.new_context(storage_state=estado) if estado \
                      else await browser.new_context()
            except Exception:
                ctx = await browser.new_context()
            page = await ctx.new_page()
            if estado:                  # sesión recordada: la ventana no hace falta
                await self._nav_minimizar(ctx, page)

            # ¿Sigue viva la sesión recordada? (sondeo a la API; si no, login manual)
            tok = uid = ""
            hdrs = None
            viva = False
            if estado:
                cookies = {c["name"]: c["value"]
                           for c in await ctx.cookies("https://www.guardiandelaproductividad.com")}
                tok, uid = self._tok_uid(cookies)
                if tok and uid.isdigit():
                    hdrs = {"Authorization": f"Bearer {tok}", "Content-Type": "application/json"}
                    # Sondeo con el primer caso real (solo consulta, no escribe)
                    _c0 = str(planes[0][3][0][1]).split(".")[0].strip()
                    _s0 = str(planes[0][3][0][2]).split(".")[0].strip()
                    try:
                        r = await ctx.request.get(
                            f"{GUARDIAN_API}/plan-trabajo-anual/api/v1/cronograma-actividades-proveedor",
                            params={"scheduleNumber": _c0, "sequenceNumber": _s0, "userId": uid,
                                    "size": 1, "page": 0, "sortDirection": "asc"},
                            headers=hdrs, timeout=15000)
                        viva = r.status == 200
                    except Exception:
                        viva = False
                if viva:
                    self.app.log(self, "Sesión de GUARDIÁN recordada — sin login manual.", "ok")
                else:
                    self.app.log(self, "La sesión guardada expiró — inicia sesión de nuevo.", "info")

            if not viva:
                await self._nav_mostrar(ctx, page)      # el login manual sí necesita la ventana
                await page.goto("https://www.guardiandelaproductividad.com/login/signin",
                                timeout=60000, wait_until="domcontentloaded")
                ev = threading.Event()
                self.after(0, lambda: self._ask_ready(ev, "Guardián (solo inicia sesión — el bot consulta por API)"))
                ev.wait()
                if self._stop:
                    await browser.close(); self.after(0, self._on_finish); return

                # Token y usuario desde las cookies de la sesión iniciada
                cookies = {c["name"]: c["value"]
                           for c in await ctx.cookies("https://www.guardiandelaproductividad.com")}
                tok, uid = self._tok_uid(cookies)
                if not uid.isdigit():
                    self.app.log(self, f"user_id no reconocido: {uid!r}", "error")
                    await browser.close(); self.after(0, self._on_finish); return
                if not tok or not uid:
                    self.app.log(self, "No se encontró la sesión — ¿iniciaste sesión?", "error")
                    await browser.close(); self.after(0, self._on_finish); return
                hdrs = {"Authorization": f"Bearer {tok}", "Content-Type": "application/json"}
                try: self._nav_guardar(await ctx.storage_state())
                except Exception: pass
                await self._nav_minimizar(ctx, page)    # login listo: fuera del medio

            self.app.log(self, f"Sesión OK (usuario {uid}) — consultando por API…", "ok")

            corte = False
            for nombre, ws, out_c, jobs in planes:
                if self._stop or corte: break
                if multi:
                    self.app.log(self, f"— Hoja «{nombre}» ({len(jobs)} filas) —", "info")
                for rn, cron, sec in jobs:
                    if self._stop: break
                    hecho += 1
                    pre = f"[{nombre}] " if multi else ""
                    self.after(0, lambda h=hecho, c=cron, s=sec, p=pre:
                        self._upd(h, total, f"{p}Cron {c} / Sec {s}")
                    )
                    try:
                        result = await self._procesar_fila_api(ctx, hdrs, uid, cron, sec)
                    except Exception as e:
                        if "Sesión expirada" in str(e):
                            self.app.log(self, "Sesión expirada — proceso detenido. "
                                               "Vuelve a iniciar sesión y reinicia.", "error")
                            hecho -= 1
                            corte = True
                            break
                        result = f"ERROR: {str(e)[:80]}"
                        self.app.log(self, result, "error")

                    self._res_add(*self._categoria(result))
                    cell = ws.cell(row=rn, column=out_c)
                    cell.value = result
                    self._autosave_dirty = True
                    ru = result.upper()
                    if "RECHAZADO" in ru:
                        self.pintar(cell, "guardian", "rechazado"); lvl = "error"
                    elif "NO REQUIERE" in ru:
                        self.pintar(cell, "guardian", "noreq");     lvl = "ok"
                    elif "PENDIENTE" in ru:
                        self.pintar(cell, "guardian", "pendiente"); lvl = "warn"
                    elif "APROBADO" in ru:
                        self.pintar(cell, "guardian", "aprobado");  lvl = "ok"
                    elif "NO HAY INFORME" in ru or "NO NECESITA" in ru:
                        self.pintar(cell, "guardian", "noinf");     lvl = "warn"
                    elif "NO ENCONTR" in ru or "ACTIVIDAD NO" in ru:
                        self.pintar(cell, "guardian", "noenc")
                        no_resultado.append(f"{cron}/{sec}"); lvl = "error"
                    else:
                        self.pintar(cell, "guardian", "otros")
                        lvl = "orange"   # estado desconocido → visible, nunca gris

                    self.app.log(self, f"{cron}/{sec}: {result}", lvl)
                    if hecho % self.AUTOSAVE_CADA == 0:
                        self._autosave(force=False)

            self._autosave()
            try: self._nav_guardar(await ctx.storage_state())
            except Exception: pass
            await browser.close()

        if no_resultado:
            self.app.log(self, f"Sin resultado ({len(no_resultado)}): {', '.join(no_resultado)}", "warn")
        self._res_fin(hecho)
        self.app.log(self, f"Listo. {hecho}/{total} procesados.", "ok")
        self.after(0, self._on_finish)

    # ── Lógica de una fila (API directa — descubierta del sitio real) ───────
    # 1) GET  /plan-trabajo-anual/api/v1/cronograma-actividades-proveedor
    #         ?scheduleNumber&sequenceNumber&userId&size&page&sortDirection
    #    → data.data[0] = {id, scheduleId, ...}
    # 2) POST /comunes/api/v1/consultaranexos
    #         {type, parentModelId=scheduleId, modelId=id, createdBy=userId, ...}
    #    → body.rows[] = {approvalStatus:{value,title}, createdAt}
    async def _procesar_fila_api(self, ctx, hdrs, uid, cronograma, secuencia):
        cron = str(cronograma).split(".")[0].strip()
        sec  = str(secuencia).split(".")[0].strip()

        r = await ctx.request.get(
            f"{GUARDIAN_API}/plan-trabajo-anual/api/v1/cronograma-actividades-proveedor",
            params={"scheduleNumber": cron, "sequenceNumber": sec, "userId": uid,
                    "size": 5, "page": 0, "sortDirection": "asc"},
            headers=hdrs, timeout=30000)
        if r.status in (401, 403):
            raise RuntimeError("Sesión expirada")
        j = await r.json()
        acts = ((j.get("data") or {}).get("data")) or []
        if not acts:
            return "ACTIVIDAD NO ENCONTRADA"
        act = acts[0]

        r2 = await ctx.request.post(
            f"{GUARDIAN_API}/comunes/api/v1/consultaranexos",
            data=json.dumps({"type": "plan_trabajo_anual",
                             "parentModelId": str(act["scheduleId"]),
                             "modelId": str(act["id"]),
                             "size": 50, "page": 0,
                             "createdBy": int(uid),
                             "sortField": "approvalStatus",
                             "sortDirection": "",
                             "isProvider": True}),
            headers=hdrs, timeout=30000)
        if r2.status in (401, 403):
            raise RuntimeError("Sesión expirada")
        j2 = await r2.json()
        rows = ((j2.get("body") or {}).get("rows")) or []
        if not rows:
            # La actividad existe pero no tiene documentos ni nada cargado
            return "NO HAY INFORME (REVISAR)"

        # documento más reciente por fecha de carga
        rows.sort(key=lambda d: str(d.get("createdAt") or ""), reverse=True)
        title = str(((rows[0].get("approvalStatus") or {}).get("title")) or "").strip()
        if not title:
            return "NO HAY INFORME (REVISAR)"
        return title.upper()  # texto tal cual lo devuelve Guardián, en mayúsculas
